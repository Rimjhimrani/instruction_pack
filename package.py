import streamlit as st
import pandas as pd
import numpy as np
import os
import json
import hashlib
import tempfile
import shutil
from pathlib import Path
from collections import defaultdict
import zipfile
from PIL import Image
import base64
import traceback
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as OpenpyxlImage
import re
from datetime import datetime
from difflib import SequenceMatcher

def navigate_to_step(step_number):
    """Helper function to navigate between steps"""
    if 1 <= step_number <= 6:
        st.session_state.current_step = step_number
        st.rerun()

# Configure Streamlit page
st.set_page_config(
    page_title="AI Packaging Template Mapper",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)
# Initialize session state
if 'current_step' not in st.session_state:
    st.session_state.current_step = 1
if 'selected_packaging_type' not in st.session_state:
    st.session_state.selected_packaging_type = None
if 'template_file' not in st.session_state:
    st.session_state.template_file = None
if 'data_file' not in st.session_state:
    st.session_state.data_file = None
if 'mapping_completed' not in st.session_state:
    st.session_state.mapping_completed = False
if 'image_option' not in st.session_state:
    st.session_state.image_option = None
if 'uploaded_images' not in st.session_state:
    st.session_state.uploaded_images = {}
if 'extracted_excel_images' not in st.session_state:
    st.session_state.extracted_excel_images = {}
if 'all_row_data' not in st.session_state:
    st.session_state.all_row_data = []

def navigate_to_step(step):
    st.session_state.current_step = step
    st.rerun()


class EnhancedImageExtractor:
    """Advanced image extraction and placement with smart positioning"""
    
    def __init__(self):
        self.supported_formats = ['.png', '.jpg', '.jpeg', '.gif', '.bmp']
        self._placement_counters = defaultdict(int)
        self.current_excel_path = None
        self.row_image_mapping = {}  # Store mapping of rows to images
        
    def analyze_template_structure(self, template_path):
        """Analyze template to find image placement areas"""
        try:
            workbook = openpyxl.load_workbook(template_path, data_only=False)
            worksheet = workbook.active
            
            image_zones = {
                'current_packaging': None,
                'primary_packaging': None,
                'secondary_packaging': None,
                'label': None
            }
            
            # Search for specific headers/keywords in the template
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_value_lower = cell.value.lower()
                        
                        # Look for packaging-related headers
                        if 'current packaging' in cell_value_lower:
                            image_zones['current_packaging'] = self._find_image_area_near_cell(worksheet, cell)
                        elif 'primary packaging' in cell_value_lower:
                            image_zones['primary_packaging'] = self._find_image_area_near_cell(worksheet, cell)
                        elif 'secondary packaging' in cell_value_lower:
                            image_zones['secondary_packaging'] = self._find_image_area_near_cell(worksheet, cell)
                        elif 'label' in cell_value_lower or 'barcode' in cell_value_lower:
                            image_zones['label'] = self._find_image_area_near_cell(worksheet, cell)
            
            workbook.close()
            return image_zones
            
        except Exception as e:
            st.error(f"Error analyzing template structure: {e}")
            return {}
    
    def _find_image_area_near_cell(self, worksheet, header_cell):
        """Find the best area for image placement near a header cell"""
        try:
            header_row = header_cell.row
            header_col = header_cell.column
            
            # Look for merged cells or empty areas below/adjacent to header
            # Strategy 1: Look directly below the header
            for row_offset in range(1, 10):  # Check up to 10 rows below
                target_row = header_row + row_offset
                target_cell = worksheet.cell(target_row, header_col)
                
                # If we find an empty area or a large merged cell, use it
                if not target_cell.value or target_cell.value == "":
                    # Check if this area has enough space (at least 3x3 cells)
                    if self._check_area_availability(worksheet, target_row, header_col, 3, 3):
                        return {
                            'cell': f"{get_column_letter(header_col)}{target_row}",
                            'row': target_row,
                            'col': header_col,
                            'width_cells': 3,
                            'height_cells': 3
                        }
            
            # Strategy 2: Look to the right of header
            for col_offset in range(1, 5):
                target_col = header_col + col_offset
                target_cell = worksheet.cell(header_row, target_col)
                
                if not target_cell.value or target_cell.value == "":
                    if self._check_area_availability(worksheet, header_row, target_col, 3, 3):
                        return {
                            'cell': f"{get_column_letter(target_col)}{header_row}",
                            'row': header_row,
                            'col': target_col,
                            'width_cells': 3,
                            'height_cells': 3
                        }
            
            # Fallback: Use a position relative to header
            return {
                'cell': f"{get_column_letter(header_col)}{header_row + 2}",
                'row': header_row + 2,
                'col': header_col,
                'width_cells': 3,
                'height_cells': 3
            }
            
        except Exception as e:
            st.warning(f"Error finding image area near cell: {e}")
            return None
    
    def _check_area_availability(self, worksheet, start_row, start_col, width, height):
        """Check if an area is available for image placement"""
        try:
            for row in range(start_row, start_row + height):
                for col in range(start_col, start_col + width):
                    cell = worksheet.cell(row, col)
                    if cell.value and str(cell.value).strip():
                        return False
            return True
        except:
            return False
    
    def extract_images_from_excel(self, excel_file_path):
        """Enhanced image extraction with better organization and row mapping"""
        try:
            self.current_excel_path = excel_file_path
            images = {}
            
            st.write("🔍 Extracting images from Excel file...")
            
            # METHOD 1: Standard openpyxl extraction with position detection
            try:
                result1 = self._extract_with_openpyxl_enhanced(excel_file_path)
                images.update(result1)
                st.write(f"✅ Enhanced extraction found {len(result1)} images")
            except Exception as e:
                st.write(f"⚠️ Enhanced extraction failed: {e}")
            
            # METHOD 2: ZIP-based extraction as fallback
            if not images:
                try:
                    result2 = self._extract_with_zipfile_enhanced(excel_file_path)
                    images.update(result2)
                    st.write(f"✅ ZIP extraction found {len(result2)} images")
                except Exception as e:
                    st.write(f"⚠️ ZIP extraction failed: {e}")
            
            if not images:
                st.warning("⚠️ No images found in Excel file.")
            else:
                st.success(f"🎯 Total images extracted: {len(images)}")
                # Build row-to-image mapping
                self._build_row_image_mapping(images)
                # Group images by suspected content
                grouped_images = self._group_images_by_content(images)
                self._display_image_groups(grouped_images)
            
            return {'all_sheets': images}
            
        except Exception as e:
            st.error(f"❌ Error extracting images: {e}")
            return {}

    def _build_row_image_mapping(self, images):
        """Build a mapping of Excel rows to their associated images"""
        self.row_image_mapping = {}
        
        # Group images by row
        for img_key, img_data in images.items():
            row_num = img_data.get('row', 0)
            if row_num > 0:  # Valid row
                if row_num not in self.row_image_mapping:
                    self.row_image_mapping[row_num] = {}
                
                img_type = img_data.get('type', 'current')
                self.row_image_mapping[row_num][img_type] = img_key
        
        st.write(f"📊 Built mapping for {len(self.row_image_mapping)} rows with images")

    def extract_images_for_part(self, data_file, part_number, all_extracted_images, vendor_code, current_row=None):
        """Extract images relevant to a specific part number - FIXED VERSION"""
        try:
            if not all_extracted_images or 'all_sheets' not in all_extracted_images:
                st.warning(f"No images available for part {part_number}")
                return self._get_fallback_images()
            
            all_images = all_extracted_images['all_sheets']
            
            # Strategy 1: Try to use row-specific mapping if available
            if current_row and current_row in self.row_image_mapping:
                row_images = {}
                for img_type, img_key in self.row_image_mapping[current_row].items():
                    if img_key in all_images:
                        row_images[img_key] = all_images[img_key]
                
                if row_images:
                    st.write(f"✅ Found {len(row_images)} row-specific images for part {part_number}")
                    return row_images
            
            # Strategy 2: Use intelligent distribution - give each part a different set
            part_specific_images = self._distribute_images_intelligently(
                all_images, part_number, vendor_code
            )
            
            if not part_specific_images:
                st.warning(f"No specific images found for part {part_number}, using fallback")
                part_specific_images = self._get_fallback_images()
            
            st.write(f"🎯 Assigned {len(part_specific_images)} images to part {part_number}")
            return part_specific_images
            
        except Exception as e:
            st.error(f"Error extracting images for part {part_number}: {e}")
            return self._get_fallback_images()
    
    def _distribute_images_intelligently(self, all_images, part_number, vendor_code):
        """Intelligently distribute images to different parts"""
        try:
            # Group images by type
            images_by_type = {}
            for img_key, img_data in all_images.items():
                img_type = img_data.get('type', 'current')
                if img_type not in images_by_type:
                    images_by_type[img_type] = []
                images_by_type[img_type].append((img_key, img_data))
            
            # Create a hash from part number to ensure consistent but different distribution
            part_hash = hash(part_number) % 1000
            
            selected_images = {}
            
            # For each type, select an image based on the part number
            for img_type, images_list in images_by_type.items():
                if images_list:
                    # Use modulo to cycle through available images
                    index = part_hash % len(images_list)
                    selected_key, selected_data = images_list[index]
                    selected_images[selected_key] = selected_data
                    
                    # Move to next image for next part (simple rotation)
                    part_hash += 1
            
            return selected_images
            
        except Exception as e:
            st.write(f"⚠️ Error in intelligent distribution: {e}")
            return self._get_first_of_each_type(all_images)
    
    def _get_fallback_images(self):
        """Provide fallback when no images are available"""
        st.write("📝 Using placeholder images (no actual images available)")
        return {
            'placeholder_current': {
                'data': None,
                'type': 'current',
                'size': (400, 300),
                'placeholder': True
            },
            'placeholder_primary': {
                'data': None,
                'type': 'primary',
                'size': (400, 300),
                'placeholder': True
            }
        }
    
    def _get_first_of_each_type(self, all_images):
        """Get the first image of each type as fallback"""
        first_images = {}
        seen_types = set()
        
        for img_key, img_data in all_images.items():
            img_type = img_data.get('type', 'current')
            if img_type not in seen_types:
                first_images[img_key] = img_data
                seen_types.add(img_type)
                if len(first_images) >= 4:  # Max 4 types
                    break
        
        return first_images
    
    def _extract_row_number(self, position):
        """Extract row number from cell position like 'A42' -> 42"""
        try:
            import re
            match = re.search(r'(\d+)', position)
            return int(match.group(1)) if match else 0
        except:
            return 0
    
    def _extract_with_openpyxl_enhanced(self, excel_file_path):
        """Enhanced openpyxl extraction with better positioning and row tracking"""
        images = {}
        
        try:
            workbook = openpyxl.load_workbook(excel_file_path, data_only=False)
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                if hasattr(worksheet, '_images') and worksheet._images:
                    for idx, img in enumerate(worksheet._images):
                        try:
                            # Get image data
                            image_data = img._data()
                            image_hash = hashlib.md5(image_data).hexdigest()
                            
                            # Create PIL Image for analysis
                            pil_image = Image.open(io.BytesIO(image_data))
                            
                            # Enhanced position detection
                            position_info = self._get_enhanced_position_info(img, worksheet, idx)
                            
                            # Smart image type classification based on COLUMN, not row
                            image_type = self._smart_classify_image_type_by_column(
                                position_info, worksheet, idx
                            )
                            
                            # Convert to base64
                            buffered = io.BytesIO()
                            pil_image.save(buffered, format="PNG")
                            img_str = base64.b64encode(buffered.getvalue()).decode()
                            
                            image_key = f"{image_type}_{sheet_name}_{position_info['position']}_{idx}"
                            images[image_key] = {
                                'data': img_str,
                                'format': 'PNG',
                                'size': pil_image.size,
                                'position': position_info['position'],
                                'row': position_info['row'],  # Important for row tracking
                                'col': position_info['col'],
                                'sheet': sheet_name,
                                'index': idx,
                                'type': image_type,
                                'hash': image_hash,
                                'confidence': position_info.get('confidence', 0.5),
                                'area_context': position_info.get('context', '')
                            }
                            
                        except Exception as e:
                            st.write(f"❌ Failed to extract image {idx} from {sheet_name}: {e}")
            
            workbook.close()
            
        except Exception as e:
            raise Exception(f"Error in enhanced openpyxl extraction: {e}")
        
        return images
    
    def _smart_classify_image_type_by_column(self, position_info, worksheet, idx):
        """Classify image type based on which COLUMN it's in, not row context"""
        try:
            col = position_info['col']
            
            # Check the column header (first few rows of this column)
            for row_num in range(1, 6):  # Check first 5 rows
                try:
                    cell = worksheet.cell(row_num, col + 1)  # +1 because openpyxl is 0-indexed
                    if cell.value and isinstance(cell.value, str):
                        header_text = cell.value.lower().strip()
                        
                        # Match based on column headers from your Excel
                        if 'current packaging' in header_text:
                            return 'current'
                        elif 'primary packaging' in header_text:
                            return 'primary'
                        elif 'secondary packaging' in header_text:
                            return 'secondary'
                        elif 'label' in header_text or 'barcode' in header_text:
                            return 'label'
                except:
                    continue
            
            # Fallback: use column position to guess type
            # Based on typical Excel layout: Current, Primary, Secondary, Label
            if col <= 5:  # Assuming first few columns
                return 'current'
            elif col <= 10:
                return 'primary' 
            elif col <= 15:
                return 'secondary'
            else:
                return 'label'
                
        except Exception as e:
            st.write(f"⚠️ Column classification failed: {e}")
            # Final fallback
            types = ['current', 'primary', 'secondary', 'label']
            return types[idx % len(types)]
    
    def _get_enhanced_position_info(self, img, worksheet, idx):
        """Get enhanced position information including context"""
        try:
            anchor = img.anchor
            if hasattr(anchor, '_from') and anchor._from:
                col = anchor._from.col
                row = anchor._from.row
                position = f"{get_column_letter(col + 1)}{row + 1}"
                
                # Analyze surrounding context (but don't rely on it for type classification)
                context = self._analyze_surrounding_context(worksheet, row, col)
                confidence = 0.8 if context else 0.5
                
                return {
                    'position': position,
                    'row': row,
                    'col': col,
                    'context': context,
                    'confidence': confidence
                }
            else:
                return {
                    'position': f"Image_{idx + 1}",
                    'row': 0,
                    'col': 0,
                    'context': '',
                    'confidence': 0.3
                }
                
        except Exception as e:
            return {
                'position': f"Unknown_{idx}",
                'row': 0,
                'col': 0,
                'context': '',
                'confidence': 0.1
            }
    
    def _analyze_surrounding_context(self, worksheet, img_row, img_col):
        """Analyze text around image position to determine its purpose"""
        context_keywords = {
            'current': ['current', 'present', 'existing', 'actual'],
            'primary': ['primary', 'main', 'first', 'initial'],
            'secondary': ['secondary', 'outer', 'external', 'second'],
            'label': ['label', 'barcode', 'sticker', 'tag']
        }
        
        # First, check column headers (more reliable for your structure)
        header_context = self._check_column_headers(worksheet, img_col)
        if header_context:
            return header_context
        
        # Fallback: Check cells in a 5x5 area around the image
        found_context = []
        for row_offset in range(-2, 3):
            for col_offset in range(-2, 3):
                try:
                    cell = worksheet.cell(img_row + row_offset, img_col + col_offset)
                    if cell.value and isinstance(cell.value, str):
                        cell_text = cell.value.lower()
                        for context_type, keywords in context_keywords.items():
                            if any(keyword in cell_text for keyword in keywords):
                                found_context.append(context_type)
                except:
                    continue
        
        return ', '.join(set(found_context)) if found_context else ''
    
    def _check_column_headers(self, worksheet, img_col):
        """Check column headers to determine image type based on your Excel structure"""
        try:
            # Check the first few rows for headers
            for row_num in range(1, 5):
                cell = worksheet.cell(row_num, img_col + 1)  # +1 for 1-based indexing
                if cell.value and isinstance(cell.value, str):
                    header_text = cell.value.lower()
                    
                    if 'current packaging' in header_text:
                        return 'current'
                    elif 'primary packaging' in header_text:
                        return 'primary'
                    elif 'secondary packaging' in header_text:
                        return 'secondary'
                    elif 'label' in header_text:
                        return 'label'
            
            return ''
        except:
            return ''
    
    def _extract_with_zipfile_enhanced(self, excel_file_path):
        """Enhanced ZIP extraction with smart organization"""
        images = {}
        
        try:
            with zipfile.ZipFile(excel_file_path, 'r') as zip_ref:
                file_list = zip_ref.namelist()
                
                # Find all image files
                image_files = []
                for f in file_list:
                    if any(f.lower().endswith(ext) for ext in self.supported_formats):
                        image_files.append(f)
                
                # Sort images to maintain consistent order
                image_files.sort()
                
                for idx, image_file in enumerate(image_files):
                    try:
                        with zip_ref.open(image_file) as img_file:
                            image_data = img_file.read()
                            
                            pil_image = Image.open(io.BytesIO(image_data))
                            
                            # Convert to base64
                            buffered = io.BytesIO()
                            pil_image.save(buffered, format="PNG")
                            img_str = base64.b64encode(buffered.getvalue()).decode()
                            
                            # Create hash
                            image_hash = hashlib.md5(image_data).hexdigest()
                            
                            # Smart type classification for ZIP-extracted images
                            filename = os.path.basename(image_file).lower()
                            image_type = self._classify_from_filename(filename, idx)
                            
                            image_key = f"{image_type}_ZIP_{filename}_{idx}"
                            
                            images[image_key] = {
                                'data': img_str,
                                'format': 'PNG',
                                'size': pil_image.size,
                                'position': f"ZIP_{idx}",
                                'row': idx + 2,  # Assign sequential rows
                                'col': 0,
                                'sheet': 'ZIP_EXTRACTED',
                                'index': idx,
                                'type': image_type,
                                'hash': image_hash,
                                'source_path': image_file,
                                'confidence': 0.6
                            }
                            
                    except Exception as e:
                        st.write(f"❌ Failed to extract {image_file}: {e}")
        
        except Exception as e:
            raise Exception(f"Error in enhanced ZIP extraction: {e}")
        
        return images
    
    def _classify_from_filename(self, filename, idx):
        """Classify image type based on filename patterns"""
        filename_keywords = {
            'current': ['current', 'present', 'actual', 'now'],
            'primary': ['primary', 'main', 'inner', 'first', '1st'],
            'secondary': ['secondary', 'outer', 'external', 'second', '2nd'],
            'label': ['label', 'barcode', 'tag', 'sticker', 'code']
        }
        
        for img_type, keywords in filename_keywords.items():
            if any(keyword in filename for keyword in keywords):
                return img_type
        
        # Default fallback - cycle through types
        types = ['current', 'primary', 'secondary', 'label']
        return types[idx % len(types)]
    
    def _group_images_by_content(self, images):
        """Group images by their classified types for better organization"""
        grouped = defaultdict(list)
        for key, img_data in images.items():
            img_type = img_data['type']
            grouped[img_type].append((key, img_data))
        return dict(grouped)
    
    def _display_image_groups(self, grouped_images):
        """Display images organized by type"""
        st.subheader("📋 Extracted Images by Type")
        
        for img_type, images_list in grouped_images.items():
            with st.expander(f"{img_type.capitalize()} Images ({len(images_list)} found)"):
                cols = st.columns(min(3, len(images_list)))
                for idx, (key, img_data) in enumerate(images_list):
                    with cols[idx % 3]:
                        st.image(
                            f"data:image/png;base64,{img_data['data']}", 
                            caption=f"{key}\nSize: {img_data['size']}\nConfidence: {img_data.get('confidence', 0.5):.1f}",
                            width=150
                        )
    
    def smart_add_images_to_template(self, template_path, worksheet, uploaded_images):
        """Smart image placement based on template analysis"""
        try:
            added_images = 0
            temp_image_paths = []
            # Analyze template structure
            image_zones = self.analyze_template_structure(template_path)
        
            # Map image types to detected zones
            type_zone_mapping = {
                'current': image_zones.get('current_packaging'),
                'primary': image_zones.get('primary_packaging'),
                'secondary': image_zones.get('secondary_packaging'),
                'label': image_zones.get('label')
            }
        
            for img_key, img_data in uploaded_images.items():
                img_type = img_data.get('type', 'current')
                target_zone = type_zone_mapping.get(img_type)
            
                # Set size based on image type - FIXED SIZE LOGIC
                if img_type == 'current':
                    width_cm, height_cm = 8.3, 8.3  # Current packaging is larger
                else:
                    width_cm, height_cm = 4.3, 4.3  # Primary, secondary, label are smaller
            
                if target_zone:
                    # Use detected zone - but update zone size based on image type
                    if img_type == 'current':
                        # Larger zone for current packaging
                        target_zone['width_cells'] = 4  # Wider for 8.3cm
                        target_zone['height_cells'] = 15  # Taller for 8.3cm
                    else:
                        # Smaller zone for other types
                        target_zone['width_cells'] = 2  # Standard for 4.3cm
                        target_zone['height_cells'] = 7  # Standard for 4.3cm
                    
                    success = self._place_image_smart(
                        worksheet, img_key, img_data, target_zone, temp_image_paths
                    )
                else:
                    # Fallback to default positions with correct sizes
                    fallback_positions = {
                        'current': 'W4',
                        'primary': 'A42',
                        'secondary': 'G42',
                        'label': 'M42'
                    }
                    position = fallback_positions.get(img_type, 'A1')
                    success = self._place_image_at_position(
                        worksheet, img_key, img_data, position,
                        width_cm, height_cm, temp_image_paths  # Use the correct size variables
                    )
            
                if success:
                    added_images += 1
                else:
                    st.write(f"⚠️ Failed to place {img_type} image")
            return added_images, temp_image_paths
        
        except Exception as e:
            st.error(f"Error in smart image placement: {e}")
            return 0, []
    
    def _place_image_smart(self, worksheet, img_key, img_data, zone_info, temp_image_paths):
        """Place image using smart zone information - FIXED VERSION"""
        try:
            # Handle placeholder images
            if img_data.get('placeholder'):
                st.write(f"⏭️ Skipping placeholder image {img_key}")
                return True
             # Create temporary image file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_img:
                image_bytes = base64.b64decode(img_data['data'])
                tmp_img.write(image_bytes)
                tmp_img_path = tmp_img.name
            # Create openpyxl image object
            img = OpenpyxlImage(tmp_img_path)
        
            # Calculate size based on zone dimensions
            cell_width_px = 80  # Approximate Excel cell width in pixels
            cell_height_px = 20  # Approximate Excel cell height in pixels
        
            img.width = zone_info['width_cells'] * cell_width_px
            img.height = zone_info['height_cells'] * cell_height_px
        
            # Set position
            img.anchor = zone_info['cell']
        
            # Add image to worksheet
            worksheet.add_image(img)
        
            # Track temporary file for cleanup
            temp_image_paths.append(tmp_img_path)
        
            return True
        except Exception as e:
            st.write(f"❌ Failed to place {img_key} in smart zone: {e}")
            # Clean up temp file if it was created
            if 'tmp_img_path' in locals():
                try:
                    os.unlink(tmp_img_path)
                except:
                    pass
            return False
        
    def add_images_to_template(self, worksheet, uploaded_images):
        """Add uploaded images to template at specific positions - ENHANCED WITH DEBUGGING"""
        try:
            added_images = 0
            temp_image_paths = []
        
            st.write(f"🔍 Starting image placement process with {len(uploaded_images)} images")
        
            # Fixed positions for different image types
            positions = {
                'current': 'W4',      # Current packaging at T3
                'primary': 'A42',     # Primary packaging at A42
                'secondary': 'G42',   # Secondary packaging at F42
                'label': 'M42'        # Label at K42
            }
        
            for img_key, img_data in uploaded_images.items():
                st.write(f"📍 Processing image: {img_key}")
            
                # Skip placeholder images
                if img_data.get('placeholder'):
                    st.write(f"⏭️ Skipping placeholder for {img_key}")
                    continue
                img_type = img_data.get('type', 'current')
                st.write(f"🏷️ Image type: {img_type}")
            
                if img_type in positions:
                    position = positions[img_type]
                    st.write(f"📍 Target position: {position}")
                
                    # Different sizes for different types
                    if img_type == 'current':
                        width_cm, height_cm = 8.3, 8.3
                    else:
                        width_cm, height_cm = 4.3, 4.3
                
                    st.write(f"📏 Image size: {width_cm}x{height_cm} cm")
                
                    success = self._place_image_at_position(
                        worksheet, img_key, img_data, position,
                        width_cm, height_cm, temp_image_paths
                    )
                
                    if success:
                        added_images += 1
                        st.write(f"✅ Successfully added {img_key} to template")
                    else:
                        st.write(f"❌ Failed to add {img_key} to template")
                else:
                    st.write(f"⚠️ Unknown image type: {img_type}")
                st.write(f"📊 Final result: {added_images} images successfully added to template")
                return added_images, temp_image_paths
        except Exception as e:
            st.error(f"❌ Error in add_images_to_template: {e}")
            st.write(f"Error details: {str(e)}")
            return 0, []

    def _place_image_at_position(self, worksheet, img_key, img_data, cell_position, width_cm, height_cm, temp_image_paths):
        """Place a single image at the specified cell position - FIXED VERSION"""
        try:
            # Skip placeholder images
            if img_data.get('placeholder'):
                st.write(f"⏭️ Skipping placeholder image {img_key}")
                return True
            # Create temporary image file
            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_img:
                image_bytes = base64.b64decode(img_data['data'])
                tmp_img.write(image_bytes)
                tmp_img_path = tmp_img.name
            # Create openpyxl image object
            img = OpenpyxlImage(tmp_img_path)
            # Set image size (converting cm to pixels: 1cm ≈ 37.8 pixels)
            img.width = int(width_cm * 37.8)
            img.height = int(height_cm * 37.8)
            # Set position using simple anchor
            img.anchor = cell_position
            # Add image to worksheet
            worksheet.add_image(img)
            # Track temporary file for cleanup
            temp_image_paths.append(tmp_img_path)
            return True
        except Exception as e:
            st.write(f"❌ Failed to place {img_key} at {cell_position}: {e}")
            # Clean up temp file if it was created
            if 'tmp_img_path' in locals():
                try:
                    os.unlink(tmp_img_path)
                except:
                    pass
            return False

            
class EnhancedTemplateMapperWithImages:
    def __init__(self):
        self.image_extractor = EnhancedImageExtractor()
        self.similarity_threshold = 0.3
        
        # Enhanced section-based mapping rules (from your working code)
        self.section_mappings = {
            'general_information': {
                'section_keywords': ['general information', 'document info', 'metadata'],
                'field_mappings': {
                    'date': 'Date',
                    'revision no': 'Revision No.'
                }
            },
            
            'primary_packaging': {
                'section_keywords': [
                    'primary packaging instruction', 'primary packaging', 'primary', 
                    'internal', '( primary / internal )', 'primary / internal'
                ],
                'field_mappings': {
                    'primary packaging type': 'Primary Packaging Type',
                    'packaging type': 'Primary Packaging Type',
                    'l-mm': 'Primary L-mm',
                    'l mm': 'Primary L-mm',
                    'length': 'Primary L-mm',
                    'w-mm': 'Primary W-mm',
                    'w mm': 'Primary W-mm', 
                    'width': 'Primary W-mm',
                    'h-mm': 'Primary H-mm',
                    'h mm': 'Primary H-mm',
                    'height': 'Primary H-mm',
                    'qty/pack': 'Primary Qty/Pack',
                    'quantity': 'Primary Qty/Pack',
                    'empty weight': 'Primary Empty Weight',
                    'pack weight': 'Primary Pack Weight'
                }
            },
            
            'secondary_packaging': {
                'section_keywords': [
                    'secondary packaging instruction', 'secondary packaging', 'secondary', 
                    'outer', 'external', '( outer / external )', 'outer / external'
                ],
                'field_mappings': {
                    'secondary packaging type': 'Secondary Packaging Type',
                    'packaging type': 'Secondary Packaging Type',
                    'type': 'Secondary Packaging Type',
                    'l-mm': 'Secondary L-mm',
                    'l mm': 'Secondary L-mm',
                    'length': 'Secondary L-mm',
                    'w-mm': 'Secondary W-mm',
                    'w mm': 'Secondary W-mm',
                    'width': 'Secondary W-mm',
                    'h-mm': 'Secondary H-mm',
                    'h mm': 'Secondary H-mm',
                    'height': 'Secondary H-mm',
                    'qty/pack': 'Secondary Qty/Pack',
                    'quantity': 'Secondary Qty/Pack',
                    'empty weight': 'Secondary Empty Weight',
                    'pack weight': 'Secondary Pack Weight'
                }
            },
            
            'part_information': {
                'section_keywords': [
                    'part information', 'part info', 'part', 'component', 'item', 'component information'
                ],
                'field_mappings': {
                    'L': 'Part L',
                    'l': 'Part L',
                    'length': 'Part L',
                    'part l': 'Part L',
                    'component l': 'Part L',
                    'W': 'Part W',
                    'w': 'Part W',
                    'width': 'Part W',
                    'part w': 'Part W',
                    'component w': 'Part W',
                    'H': 'Part H',
                    'h': 'Part H',
                    'height': 'Part H',
                    'part h': 'Part H',
                    'component h': 'Part H',
                    'part no': 'Part No',
                    'part number': 'Part No',
                    'description': 'Part Description',
                    'unit weight': 'Part Unit Weight'
                }
            },
            
            'vendor_information': {
                'section_keywords': [
                    'vendor information', 'vendor info', 'vendor', 'supplier', 'supplier information', 'supplier info'
                ],
                'field_mappings': {
                    'vendor name': 'Vendor Name',
                    'name': 'Vendor Name',
                    'supplier name': 'Vendor Name',
                    'vendor code': 'Vendor Code',
                    'supplier code': 'Vendor Code',
                    'code': 'Vendor Code',
                    'vendor location': 'Vendor Location',
                    'location': 'Vendor Location',
                    'supplier location': 'Vendor Location',
                    'address': 'Vendor Location'
                }
            },
            
            'procedure_information': {
                'section_keywords': [
                    'procedure information', 'procedure', 'packaging procedure', 'loading details',
                    'pallet information', 'pallet details', 'packaging details',
                    'loading instruction', 'packing procedure', 'palletization'
                ],
                'field_mappings': {
                    # x No. of Parts mapping (Column AH: "x No. of Parts")
                    'x no. of parts': 'x No. of Parts',
                    'x no of parts': 'x No. of Parts',
                    'x number of parts': 'x No. of Parts',
                    'no. of parts': 'x No. of Parts',
                    'no of parts': 'x No. of Parts',
                    'number of parts': 'x No. of Parts',
                    'parts': 'x No. of Parts',
                
                    # Layer mapping (Column AF: "Layer") 
                    'layer': 'Layer',
                    'layers': 'Layer',
                    'max layer': 'Layer',
                    'maximum layer': 'Layer',
                    'pallet layer': 'Layer',
                    'boxes per layer': 'Layer',
                
                    # Level mapping (Column AG: "Level")
                    'level': 'Level',
                    'levels': 'Level',
                    'max level': 'Level',
                    'maximum level': 'Level',
                    'stacking level': 'Level',
                    'pallet level': 'Level',
                
                    # Inner/Outer dimensions
                    'inner l': 'Inner L',
                    'inner w': 'Inner W', 
                    'inner h': 'Inner H',
                    'inner length': 'Inner L',
                    'inner width': 'Inner W',
                    'inner height': 'Inner H',
                    'outer l': 'Outer L',
                    'outer w': 'Outer W',
                    'outer h': 'Outer H',
                    'outer length': 'Outer L',
                    'outer width': 'Outer W',
                    'outer height': 'Outer H',
                    'inner qty/pack': 'Inner Qty/Pack',
                    'inner quantity': 'Inner Qty/Pack',
                    'inner qty': 'Inner Qty/Pack'
                }
            }
        }

    def preprocess_text(self, text):
        """Preprocess text for better matching"""
        try:
            if pd.isna(text) or text is None:
                return ""
            
            text = str(text).lower()
            text = re.sub(r'[()[\]{}]', ' ', text)
            text = re.sub(r'[^\w\s/-]', ' ', text)
            text = re.sub(r'\s+', ' ', text).strip()
            
            return text
        except Exception as e:
            st.error(f"Error in preprocess_text: {e}")
            return ""

    def is_mappable_field(self, text):
        """Enhanced field detection for packaging templates"""
        try:
            if not text or pd.isna(text):
                return False
            text = str(text).lower().strip()
            if not text:
                return False
            print(f"DEBUG is_mappable_field: Checking '{text}'")

            # Skip header-like patterns that should not be treated as fields
            header_exclusions = [
                'vendor information', 'part information', 'primary packaging', 'secondary packaging',
                'packaging instruction', 'procedure', 'steps', 'process'
            ]
            for exclusion in header_exclusions:
                if exclusion in text and 'type' not in text:
                    print(f"DEBUG: Excluding '{text}' as header")
                    return False
        
            # Define mappable field patterns for packaging templates
            mappable_patterns = [
                r'packaging\s+type', r'\btype\b',
                r'\bl[-\s]*mm\b', r'\bw[-\s]*mm\b', r'\bh[-\s]*mm\b',
                r'\bl\b', r'\bw\b', r'\bh\b',
                r'part\s+l\b', r'part\s+w\b', r'part\s+h\b',
                r'\blength\b', r'\bwidth\b', r'\bheight\b',
                r'qty[/\s]*pack', r'quantity\b', r'weight\b', r'empty\s+weight',
                r'\bcode\b', r'\bname\b', r'\bdescription\b', r'\blocation\b',
                r'part\s+no\b', r'part\s+number\b',
                r'\bdate\b',
                r'\brev(ision)?\s*no\.?\b',
                # Procedure-specific patterns
                r'\bx\s*no\.?\s*of\s*parts\b',
                r'\bx\s*no\s*of\s*parts\b',
                r'\bx\s*number\s*of\s*parts\b',
                r'\bno\.?\s*of\s*parts\b',
                r'\bnumber\s*of\s*parts\b',
                r'\bparts\s*per\s*pack\b',
                r'\bparts\s*quantity\b',
                r'\bqty\s*of\s*parts\b',
                r'\blevel\b', r'\blevels\b',
                r'\blayer\b', r'\blayers\b',
                r'\bmax\s*level\b', r'\bmaximum\s*level\b',
                r'\bmax\s*layer\b', r'\bmaximum\s*layer\b',
                r'\bstacking\s*level\b', r'\bpallet\s*level\b',
                r'\binner\s*l\b', r'\binner\s*length\b',
                r'\binner\s*w\b', r'\binner\s*width\b', 
                r'\binner\s*h\b', r'\binner\s*height\b',
                r'\binner\s*qty[/\s]*pack\b',
                r'\bouter\s*l\b', r'\bouter\s*length\b',
                r'\bouter\s*w\b', r'\bouter\s*width\b',
                r'\bouter\s*h\b', r'\bouter\s*height\b',
                r'\bpallet\b', r'\bpalletiz\w*\b'
            ]
        
            for pattern in mappable_patterns:
                if re.search(pattern, text):
                    print(f"DEBUG: '{text}' matches pattern '{pattern}'")
                    return True
        
            if text.endswith(':'):
                print(f"DEBUG: '{text}' ends with colon")
                return True
        
            print(f"DEBUG: '{text}' is NOT mappable")
            return False
        except Exception as e:
            st.error(f"Error in is_mappable_field: {e}")
            return False

    def identify_section_context(self, worksheet, row, col, max_search_rows=15):
        """Enhanced section identification with better pattern matching"""
        try:
            section_context = None
            for search_row in range(max(1, row - max_search_rows), row + 5):
                for search_col in range(max(1, col - 20), min(worksheet.max_column + 1, col + 20)):
                    try:
                        cell = worksheet.cell(row=search_row, column=search_col)
                        if cell.value:
                            cell_text = self.preprocess_text(str(cell.value))
                    
                            for section_name, section_info in self.section_mappings.items():
                                for keyword in section_info['section_keywords']:
                                    keyword_processed = self.preprocess_text(keyword)
                            
                                    if keyword_processed == cell_text or keyword_processed in cell_text or cell_text in keyword_processed:
                                        print(f"DEBUG: Found section context '{section_name}' for field at ({row}, {col}) via keyword '{keyword}'")
                                        return section_name
                            
                                # Enhanced context matching
                                if section_name == 'procedure_information':
                                    procedure_indicators = [
                                        'procedure', 'loading', 'pallet', 'packaging procedure',
                                        'stacking', 'palletization', 'loading details', 
                                        'packing instruction', 'step', 'layer', 'level'
                                    ]
                                    if any(indicator in cell_text for indicator in procedure_indicators):
                                        print(f"DEBUG: Found procedure context for field at ({row}, {col}) via indicator in '{cell_text}'")
                                        return section_name
                            
                                elif section_name == 'primary_packaging':
                                    if ('primary' in cell_text and ('packaging' in cell_text or 'internal' in cell_text)):
                                        return section_name
                                elif section_name == 'secondary_packaging':
                                    if ('secondary' in cell_text and ('packaging' in cell_text or 'outer' in cell_text or 'external' in cell_text)):
                                        return section_name
                                elif section_name == 'part_information':
                                    if (('part' in cell_text and ('information' in cell_text or 'info' in cell_text)) or ('component' in cell_text and ('information' in cell_text or 'info' in cell_text))):
                                        return section_name
                                elif section_name == 'vendor_information':
                                    if (('vendor' in cell_text and ('information' in cell_text or 'info' in cell_text)) or ('supplier' in cell_text and ('information' in cell_text or 'info' in cell_text))):
                                        return section_name
                    except:
                        continue
    
            return self.infer_section_from_field_name(row, col)
    
        except Exception as e:
            st.error(f"Error in identify_section_context: {e}")
            return None
        
    def infer_section_from_field_name(self, row, col):
        """Infer section context from field name when no explicit section header found"""
        try:
            return 'procedure_information'
        except Exception as e:
            return 'procedure_information'

    def calculate_similarity(self, text1, text2):
        """Calculate similarity between two texts"""
        try:
            if not text1 or not text2:
                return 0.0
            
            text1 = self.preprocess_text(text1)
            text2 = self.preprocess_text(text2)
            
            if not text1 or not text2:
                return 0.0
            
            sequence_sim = SequenceMatcher(None, text1, text2).ratio()
            return sequence_sim
        except Exception as e:
            st.error(f"Error in calculate_similarity: {e}")
            return 0.0

    def find_template_fields_with_context_and_images(self, template_file):
        """Find template fields and image upload areas"""
        fields = {}
        image_areas = []
        try:
            workbook = openpyxl.load_workbook(template_file)
            worksheet = workbook.active
        
            merged_ranges = worksheet.merged_cells.ranges
        
            for row in worksheet.iter_rows():
                for cell in row:
                    try:
                        if cell.value is not None:
                            cell_value = str(cell.value).strip()

                            # Force capture Date & Revision No anywhere in sheet
                            if cell_value.lower() in ['date', 'revision no.', 'revision no']:
                                fields[cell.coordinate] = {
                                    'value': cell_value,
                                    'row': cell.row,
                                    'column': cell.column,
                                    'merged_range': None,
                                    'section_context': 'general_information',
                                    'is_mappable': True
                                }
                                continue
                        
                            if cell_value and self.is_mappable_field(cell_value):
                                cell_coord = cell.coordinate
                                merged_range = None
                            
                                for merge_range in merged_ranges:
                                    if cell.coordinate in merge_range:
                                        merged_range = str(merge_range)
                                        break
                            
                                section_context = self.identify_section_context(
                                    worksheet, cell.row, cell.column
                                )
                            
                                print(f"DEBUG: Found field '{cell_value}' at {cell_coord}")
                                print(f"DEBUG: Section context: {section_context}")
                                print("---")
                            
                                fields[cell_coord] = {
                                    'value': cell_value,
                                    'row': cell.row,
                                    'column': cell.column,
                                    'merged_range': merged_range,
                                    'section_context': section_context,
                                    'is_mappable': True
                                }
                    except Exception as e:
                        continue
        
            workbook.close()
        
        except Exception as e:
            st.error(f"Error reading template: {e}")
    
        return fields, image_areas

    def find_data_cell_for_label(self, worksheet, field_info):
        """Find data cell for a label with improved merged cell handling"""
        try:
            row = field_info['row']
            col = field_info['column']
            merged_ranges = list(worksheet.merged_cells.ranges)
        
            def is_suitable_data_cell(cell_coord):
                """Check if a cell is suitable for data entry"""
                try:
                    cell = worksheet[cell_coord]
                    if hasattr(cell, '__class__') and cell.__class__.__name__ == 'MergedCell':
                        return False
                    if cell.value is None or str(cell.value).strip() == "":
                        return True
                    cell_text = str(cell.value).lower().strip()
                    data_patterns = [r'^_+$', r'^\.*$', r'^-+$', r'enter', r'fill', r'data']
                    return any(re.search(pattern, cell_text) for pattern in data_patterns)
                except:
                    return False
            
            # Strategy 1: Look right of label (most common pattern)
            for offset in range(1, 6):
                target_col = col + offset
                if target_col <= worksheet.max_column:
                    cell_coord = worksheet.cell(row=row, column=target_col).coordinate
                    if is_suitable_data_cell(cell_coord):
                        return cell_coord
            
            # Strategy 2: Look below label
            for offset in range(1, 4):
                target_row = row + offset
                if target_row <= worksheet.max_row:
                    cell_coord = worksheet.cell(row=target_row, column=col).coordinate
                    if is_suitable_data_cell(cell_coord):
                        return cell_coord
            
            # Strategy 3: Look in nearby area
            for r_offset in range(-1, 3):
                for c_offset in range(-1, 6):
                    if r_offset == 0 and c_offset == 0:
                        continue
                    target_row = row + r_offset
                    target_col = col + c_offset
                
                    if (target_row > 0 and target_row <= worksheet.max_row and 
                        target_col > 0 and target_col <= worksheet.max_column):
                            cell_coord = worksheet.cell(row=target_row, column=target_col).coordinate
                            if is_suitable_data_cell(cell_coord):
                                return cell_coord
            
            return None
            
        except Exception as e:
            st.error(f"Error in find_data_cell_for_label: {e}")
            return None

    # *** NEW METHOD: Read procedure steps from Excel template ***
    def read_procedure_steps_from_template(self, template_path, packaging_type=None):
        """
        Read procedure steps directly from the Excel template.
        Args:
            template_path: Path to the Excel templat
            packaging_type: Optional packaging type to filter steps
        Returns:
            List of procedure steps with {placeholders}
        """
        try:
            print(f"\n=== READING PROCEDURE STEPS FROM TEMPLATE ===")
            st.write(f"📖 Reading procedure steps from template...")
        
            workbook = openpyxl.load_workbook(template_path)
            worksheet = workbook.active
        
            procedure_steps = []
            start_row = 28  # Based on your original code
            target_cols = list(range(2, 19))  # Columns B to P (2 to 16)
            max_search_rows = 50  # Search up to 50 rows
            empty_count = 0       # track consecutive empty rows
            
            for row_num in range(start_row, start_row + max_search_rows):
                try:
                    row_has_content = False
                    step_text = ""
                    # Look for content in columns B to P
                    for col_num in target_cols:
                        try:
                            cell = worksheet.cell(row=row_num, column=col_num)
                            if cell.value and str(cell.value).strip():
                                cell_text = str(cell.value).strip()
                                if cell_text and not cell_text.isspace():
                                    step_text = cell_text
                                    row_has_content = True
                                    break
                        except:
                            continue
                    if row_has_content and step_text:
                        # Clean and validate the step text
                        step_text = step_text.strip()
                    
                        # Skip obviously non-procedure content
                        skip_patterns = [
                            r'^[0-9]+$',  # Just numbers
                            r'^[A-Z]$',   # Single letters
                            r'^[-_=]+$',  # Just separators
                        ]
                    
                        should_skip = any(re.match(pattern, step_text) for pattern in skip_patterns)
                    
                        if not should_skip and len(step_text) > 5:  # Minimum length check
                            procedure_steps.append(step_text)
                            print(f"📝 Found step {len(procedure_steps)}: {step_text[:50]}...")
                        empty_count = 0  # reset empty counter after a valid step
                    else:
                        empty_count += 1
                        if empty_count >= 3:  # stop after 3 consecutive empty rows
                            break
                except Exception as e:
                    print(f"⚠️ Error reading row {row_num}: {e}")
                    continue
            workbook.close()
        
            print(f"✅ Successfully read {len(procedure_steps)} procedure steps from template")
            st.write(f"✅ Found {len(procedure_steps)} procedure steps in template")
        
            # Debug: Show found steps
            for i, step in enumerate(procedure_steps, 1):
                print(f"  Step {i}: {step[:100]}...")
            return procedure_steps
        except Exception as e:
            print(f"❌ Error reading procedure steps from template: {e}")
            st.error(f"Error reading procedure steps from template: {e}")
            return []

    def substitute_placeholders_in_steps(self, procedure_steps, data_dict):
        """
        Replace placeholders in procedure steps with actual data values.
        
        Args:
            procedure_steps: List of steps with {placeholders}
            data_dict: Dictionary containing mapped data values
            
        Returns:
            List of procedure steps with placeholders replaced
        """
        try:
            print(f"\n=== SUBSTITUTING PLACEHOLDERS IN STEPS ===")
            st.write(f"🔄 Replacing placeholders with actual data...")
            
            # Debug: Print available data
            print(f"Available data in data_dict:")
            for key, value in data_dict.items():
                print(f"  '{key}': '{value}'")
            print("=" * 50)
            
            filled_steps = []
            
            for i, step in enumerate(procedure_steps, 1):
                filled_step = step
                
                print(f"Processing step {i}: {step[:50]}...")
                
                # Enhanced mapping with multiple fallback options
                replacements = {
                    # *** CRITICAL: Enhanced quantity mappings - multiple fallbacks ***
                    '{x No. of Parts}': (
                        data_dict.get('x No. of Parts') or 
                        data_dict.get('X No. of Parts') or
                        data_dict.get('x no. of parts') or
                        data_dict.get('X no. of parts') or
                        data_dict.get('no. of parts') or
                        data_dict.get('No. of Parts') or
                        data_dict.get('number of parts') or
                        data_dict.get('Number of Parts') or
                        data_dict.get('parts per pack') or
                        data_dict.get('Parts Per Pack') or
                        data_dict.get('qty of parts') or
                        data_dict.get('Qty of Parts') or
                        '8'  # Default fallback
                    ),
                
                    # *** CRITICAL: Enhanced Level mappings - multiple fallbacks ***
                    '{Level}': (
                        data_dict.get('Level') or
                        data_dict.get('level') or
                        data_dict.get('LEVEL') or
                        data_dict.get('Levels') or
                        data_dict.get('levels') or
                        data_dict.get('max level') or
                        data_dict.get('Max Level') or
                        data_dict.get('maximum level') or
                        data_dict.get('Maximum Level') or
                        data_dict.get('stacking level') or
                        data_dict.get('Stacking Level') or
                        '5'  # Default fallback
                    ),
                
                    # *** CRITICAL: Enhanced Layer mappings - multiple fallbacks ***
                    '{Layer}': (
                        data_dict.get('Layer') or
                        data_dict.get('layer') or
                        data_dict.get('LAYER') or
                        data_dict.get('Layers') or
                        data_dict.get('layers') or
                        data_dict.get('max layer') or
                        data_dict.get('Max Layer') or
                        data_dict.get('maximum layer') or
                        data_dict.get('Maximum Layer') or
                        '4'  # Default fallback
                    ),
                    
                    # Inner dimensions - try multiple key variations
                    '{Inner L}': (
                        data_dict.get('Inner L') or 
                        data_dict.get('inner l') or
                        data_dict.get('Inner l') or
                        data_dict.get('INNER L') or
                        data_dict.get('Inner Length') or
                        data_dict.get('inner length') or
                        'XXX'
                    ),
                    '{Inner W}': (
                        data_dict.get('Inner W') or 
                        data_dict.get('inner w') or
                        data_dict.get('Inner w') or
                        data_dict.get('INNER W') or
                        data_dict.get('Inner Width') or
                        data_dict.get('inner width') or
                        'XXX'
                    ),
                    '{Inner H}': (
                        data_dict.get('Inner H') or 
                        data_dict.get('inner h') or
                        data_dict.get('Inner h') or
                        data_dict.get('INNER H') or
                        data_dict.get('Inner Height') or
                        data_dict.get('inner height') or
                        'XXX'
                    ),
                    
                    # Inner Qty/Pack - try multiple variations
                    '{Inner Qty/Pack}': (
                        data_dict.get('Inner Qty/Pack') or
                        data_dict.get('inner qty/pack') or
                        data_dict.get('Inner qty/pack') or
                        data_dict.get('INNER QTY/PACK') or
                        data_dict.get('Inner Quantity') or
                        data_dict.get('inner quantity') or
                        '1'
                    ),
                    
                    # Outer dimensions - try multiple variations
                    '{Outer L}': (
                        data_dict.get('Outer L') or 
                        data_dict.get('outer l') or
                        data_dict.get('Outer l') or
                        data_dict.get('OUTER L') or
                        data_dict.get('Outer Length') or
                        data_dict.get('outer length') or
                        'XXX'
                    ),
                    '{Outer W}': (
                        data_dict.get('Outer W') or 
                        data_dict.get('outer w') or
                        data_dict.get('Outer w') or
                        data_dict.get('OUTER W') or
                        data_dict.get('Outer Width') or
                        data_dict.get('outer width') or
                        'XXX'
                    ),
                    '{Outer H}': (
                        data_dict.get('Outer H') or 
                        data_dict.get('outer h') or
                        data_dict.get('Outer h') or
                        data_dict.get('OUTER H') or
                        data_dict.get('Outer Height') or
                        data_dict.get('outer height') or
                        'XXX'
                    ),
                    
                    # Primary Qty/Pack - try multiple variations
                    '{Primary Qty/Pack}': (
                        data_dict.get('Primary Qty/Pack') or
                        data_dict.get('primary qty/pack') or
                        data_dict.get('Primary qty/pack') or
                        data_dict.get('PRIMARY QTY/PACK') or
                        data_dict.get('Primary Quantity') or
                        data_dict.get('primary quantity') or
                        '1'
                    ),
                    
                    # Generic Qty/Pack - try multiple variations
                    '{Qty/Pack}': (
                        data_dict.get('Qty/Pack') or
                        data_dict.get('qty/pack') or
                        data_dict.get('QTY/PACK') or
                        data_dict.get('Quantity') or
                        data_dict.get('quantity') or
                        '1'
                    ),
                    '{Qty/Veh}': (
                        data_dict.get('Qty/Veh') or
                        data_dict.get('qty/veh') or
                        data_dict.get('QTY/VEH') or
                        data_dict.get('Qty/Pack') or
                        data_dict.get('qty/pack') or
                        '1'
                    ),
                    
                    # Secondary dimensions
                    '{Secondary L-mm}': (
                        data_dict.get('Secondary L-mm') or
                        data_dict.get('secondary l-mm') or
                        data_dict.get('Secondary L') or
                        data_dict.get('secondary l') or
                        'XXX'
                    ),
                    '{Secondary W-mm}': (
                        data_dict.get('Secondary W-mm') or
                        data_dict.get('secondary w-mm') or
                        data_dict.get('Secondary W') or
                        data_dict.get('secondary w') or
                        'XXX'
                    ),
                    '{Secondary H-mm}': (
                        data_dict.get('Secondary H-mm') or
                        data_dict.get('secondary h-mm') or
                        data_dict.get('Secondary H') or
                        data_dict.get('secondary h') or
                        'XXX'
                    ),
                    
                    # Primary dimensions
                    '{Primary L-mm}': (
                        data_dict.get('Primary L-mm') or
                        data_dict.get('primary l-mm') or
                        data_dict.get('Primary L') or
                        data_dict.get('primary l') or
                        'XXX'
                    ),
                    '{Primary W-mm}': (
                        data_dict.get('Primary W-mm') or
                        data_dict.get('primary w-mm') or
                        data_dict.get('Primary W') or
                        data_dict.get('primary w') or
                        'XXX'
                    ),
                    '{Primary H-mm}': (
                        data_dict.get('Primary H-mm') or
                        data_dict.get('primary h-mm') or
                        data_dict.get('Primary H') or
                        data_dict.get('primary h') or
                        'XXX'
                    )
                }
                
                # Debug: Show what replacements are being made
                for placeholder, raw_value in replacements.items():
                    if placeholder in filled_step:
                        clean_value = self.clean_data_value(raw_value)
                        if not clean_value or clean_value == "":
                            clean_value = 'XXX'
                        print(f"  Replacing {placeholder} with '{clean_value}' (from: {raw_value})")
                        filled_step = filled_step.replace(placeholder, str(clean_value))
                
                filled_steps.append(filled_step)
                print(f"  Final step {i}: {filled_step[:100]}...")
                print("---")
            
            print(f"✅ Successfully processed {len(filled_steps)} procedure steps")
            st.write(f"✅ Replaced placeholders in {len(filled_steps)} steps")
            
            return filled_steps
            
        except Exception as e:
            print(f"❌ Error substituting placeholders: {e}")
            st.error(f"Error substituting placeholders: {e}")
            return procedure_steps  # Return original steps if substitution fails

    def map_data_with_section_context(self, template_fields, data_df):
        """Enhanced mapping with EXACT column name matching"""
        mapping_results = {}
        used_columns = set()

        try:
            data_columns = data_df.columns.tolist()
            print(f"DEBUG: Available data columns: {data_columns}")
        
            for coord, field in template_fields.items():
                try:
                    best_match = None
                    best_score = 0.0
                    field_value = field['value']
                    section_context = field.get('section_context')

                    print(f"DEBUG: Mapping field '{field_value}' with section '{section_context}'")

                    # Direct exact column name matching first
                    field_lower = self.preprocess_text(field_value)
                
                    for data_col in data_columns:
                        if data_col in used_columns:
                            continue
                        
                        col_lower = self.preprocess_text(data_col)
                    
                        # Exact matches
                        if col_lower == field_lower:
                            best_match = data_col
                            best_score = 1.0
                            print(f"DEBUG: DIRECT EXACT MATCH: '{field_value}' → '{data_col}'")
                            break
                    
                        # Special case matches for specific columns
                        if field_lower == 'layer' and col_lower == 'layer':
                            best_match = data_col
                            best_score = 1.0
                            print(f"DEBUG: DIRECT LAYER MATCH: '{field_value}' → '{data_col}'")
                            break
                        elif field_lower == 'level' and col_lower == 'level':
                            best_match = data_col  
                            best_score = 1.0
                            print(f"DEBUG: DIRECT LEVEL MATCH: '{field_value}' → '{data_col}'")
                            break
                        elif ('x no' in field_lower or 'no. of parts' in field_lower) and 'x no of parts' in col_lower:
                            best_match = data_col
                            best_score = 1.0
                            print(f"DEBUG: DIRECT X NO OF PARTS MATCH: '{field_value}' → '{data_col}'")
                            break

                    # If direct match found, use it
                    if best_match:
                        mapping_results[coord] = {
                            'template_field': field_value,
                            'data_column': best_match,
                            'similarity': best_score,
                            'field_info': field,
                            'section_context': section_context,
                            'is_mappable': True
                        }
                        used_columns.add(best_match)
                        print(f"DEBUG: DIRECT MATCH SUCCESS: {field_value} → {best_match}")
                        continue

                    # Force procedure context for specific fields
                    if not section_context:
                        procedure_fields = ['layer', 'level', 'x no of parts', 'no. of parts', 'parts']
                        if any(proc_field in field_lower for proc_field in procedure_fields):
                            section_context = 'procedure_information'
                            print(f"DEBUG: FORCED procedure context for field '{field_value}'")

                    # Section mapping logic
                    if section_context and section_context in self.section_mappings:
                        section_mappings = self.section_mappings[section_context]['field_mappings']
                        print(f"DEBUG: Section mappings: {section_mappings}")

                        for template_field_key, data_column_pattern in section_mappings.items():
                            normalized_field_value = self.preprocess_text(field_value)
                            normalized_template_key = self.preprocess_text(template_field_key)

                            print(f"DEBUG: Comparing '{normalized_field_value}' with '{normalized_template_key}'")

                            if normalized_field_value == normalized_template_key:
                                # For procedure_information, don't add section prefix
                                if section_context == "procedure_information":
                                    expected_column = data_column_pattern 
                                else:
                                    section_prefix = section_context.split('_')[0].capitalize()
                                    expected_column = f"{section_prefix} {data_column_pattern}".strip()
                            
                                print(f"DEBUG: Looking for expected column: '{expected_column}'")

                                for data_col in data_columns:
                                    if data_col in used_columns:
                                        continue
                                    if self.preprocess_text(data_col) == self.preprocess_text(expected_column):
                                        best_match = data_col
                                        best_score = 1.0
                                        print(f"DEBUG: SECTION EXACT MATCH FOUND: {data_col}")
                                        break

                                # Fallback to similarity match if no exact match
                                if not best_match:
                                    for data_col in data_columns:
                                        if data_col in used_columns:
                                            continue
                                        similarity = self.calculate_similarity(expected_column, data_col)
                                        if similarity > best_score and similarity >= self.similarity_threshold:
                                            best_score = similarity
                                            best_match = data_col
                                            print(f"DEBUG: SECTION SIMILARITY MATCH: {data_col} (score: {similarity})")
                                break

                    # Final fallback: general similarity matching
                    if not best_match:
                        for data_col in data_columns:
                            if data_col in used_columns:
                                continue
                            similarity = self.calculate_similarity(field_value, data_col)
                            if similarity > best_score and similarity >= self.similarity_threshold:
                                best_score = similarity
                                best_match = data_col

                    print(f"DEBUG: Final mapping result - Field: '{field_value}' → Column: '{best_match}' (Score: {best_score})")
                    print("=" * 50)

                    # Save mapping
                    mapping_results[coord] = {
                        'template_field': field_value,
                        'data_column': best_match,
                        'similarity': best_score,
                        'field_info': field,
                        'section_context': section_context,
                        'is_mappable': best_match is not None
                    }

                    if best_match:
                        used_columns.add(best_match)

                except Exception as e:
                    st.error(f"Error mapping field {coord}: {e}")
                    continue

        except Exception as e:
            st.error(f"Error in map_data_with_section_context: {e}")
        return mapping_results

    def clean_data_value(self, value):
        """Clean data value to handle NaN, None, and empty values"""
        if pd.isna(value) or value is None:
            return ""
        
        str_value = str(value).strip()
        
        if str_value.lower() in ['nan', 'none', 'null', 'n/a', '#n/a', '']:
            return ""
            
        return str_value

    def map_template_with_data(self, template_path, data_path):
        """Enhanced mapping with section-based approach and multiple row processing"""
        try:
            # Read data from Excel with proper NaN handling
            data_df = pd.read_excel(data_path)
            data_df = data_df.fillna("")
            st.write(f"📊 Loaded data with {len(data_df)} rows and {len(data_df.columns)} columns")
            
            # Force direct capture of critical procedure columns if present in Excel
            critical_cols = {
                "Outer L": ["outer l", "outer length", "outer l-mm"],
                "Outer W": ["outer w", "outer width", "outer w-mm"],
                "Outer H": ["outer h", "outer height", "outer h-mm"],
                "Inner L": ["inner l", "inner length", "inner l-mm"],
                "Inner W": ["inner w", "inner width", "inner w-mm"],
                "Inner H": ["inner h", "inner height", "inner h-mm"],
                "Primary Qty/Pack": ["Primary qty/pack", "primary qty/pack", "PRIMARY QTY/PACK"],
                "Layer":   ["layer", "layers"],
                "Level":   ["level", "levels"],
                "x No. of Parts": ["x no of parts", "x no. of parts", "x number of parts", "no. of parts", "number of parts"]
            }
            
            col_map = {}
            for canonical, variants in critical_cols.items():
                for col in data_df.columns:
                    col_norm = self.preprocess_text(col)
                    if any(col_norm == self.preprocess_text(v) for v in variants):
                        col_map[col_norm] = canonical
                        print(f"DEBUG: Matched column '{col}' ({col_norm}) -> '{canonical}'")
                        break
            
            # *** NEW: Read procedure steps from template ONCE ***
            template_procedure_steps = self.read_procedure_steps_from_template(template_path)
            if not template_procedure_steps:
                st.warning("⚠️ No procedure steps found in template. Will use empty steps.")
            
            # Store all row data for multi-template generation
            st.session_state.all_row_data = []
    
            # Process each row
            for row_idx in range(len(data_df)):
                st.write(f"🔄 Processing row {row_idx + 1}/{len(data_df)}")
                
                # Load fresh template for each row
                workbook = openpyxl.load_workbook(template_path)
                worksheet = workbook.active
        
                # Find template fields with section context
                template_fields, _ = self.find_template_fields_with_context_and_images(template_path)
        
                # Map data with section context for current row
                mapping_results = self.map_data_with_section_context_for_row(template_fields, data_df, row_idx)
        
                # Apply mappings to template
                mapping_count = 0
                data_dict = {}  # Store mapped data for procedure generation
                filename_parts = {}  # Store parts for filename
        
                for coord, mapping in mapping_results.items():
                    if mapping['is_mappable'] and mapping['data_column']:
                        try:
                            data_col = mapping['data_column']
                            raw_value = data_df[data_col].iloc[row_idx]  # Use current row
                            data_value = self.clean_data_value(raw_value)
                    
                            # Store in data_dict for procedure generation
                            data_dict[mapping['template_field']] = data_value

                            # Force map critical fields if the column matches
                            normalized_col = self.preprocess_text(data_col)
                            if normalized_col in col_map:
                                data_dict[col_map[normalized_col]] = data_value
                    
                            # Store filename components
                            field_name_lower = mapping['template_field'].lower()
                            if any(term in field_name_lower for term in ['vendor code', 'supplier code', 'code']):
                                filename_parts['vendor_code'] = data_value
                            elif 'part' in field_name_lower and ('no' in field_name_lower or 'number' in field_name_lower):
                                filename_parts['part_no'] = data_value
                            elif 'description' in field_name_lower or 'desc' in field_name_lower:
                                filename_parts['description'] = data_value
                    
                            # Find target cell and write data
                            target_cell_coord = self.find_data_cell_for_label(worksheet, mapping['field_info'])
                    
                            if target_cell_coord and data_value:
                                target_cell = worksheet[target_cell_coord]
                                target_cell.value = data_value
                                mapping_count += 1
                        except Exception as e:
                            st.write(f"⚠️ Error processing row {row_idx + 1}, field '{mapping['template_field']}': {e}")
                
                # *** NEW: Process procedure steps from template instead of hardcoded ***
                steps_written = 0
                if template_procedure_steps:
                    # Substitute placeholders with actual data
                    filled_steps = self.substitute_placeholders_in_steps(template_procedure_steps, data_dict)
                    
                    # Write the filled steps back to template
                    steps_written = self.write_filled_steps_to_template(worksheet, filled_steps)
                else:
                    st.write("⚠️ No procedure steps to process for this row")
                
                # Generate filename
                vendor_code = filename_parts.get('vendor_code', 'Unknown')
                part_no = filename_parts.get('part_no', 'Unknown') 
                description = filename_parts.get('description', 'Unknown')
        
                # Clean filename parts
                vendor_code = re.sub(r'[^\w\-_]', '', str(vendor_code))[:10]
                part_no = re.sub(r'[^\w\-_]', '', str(part_no))[:15]
                description = re.sub(r'[^\w\-_]', '', str(description))[:20]
        
                filename = f"{vendor_code}_{part_no}_{description}.xlsx"
        
                # Save workbook to temporary file
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    workbook.save(tmp_file.name)
            
                    # Store row data
                    row_data = {
                        'row_index': row_idx,
                        'filename': filename,
                        'file_path': tmp_file.name,
                        'data_dict': data_dict,
                        'mapping_count': mapping_count,
                        'steps_written': steps_written,
                        'vendor_code': vendor_code,
                        'part_no': part_no,
                        'description': description,
                        'procedure_steps': filled_steps if template_procedure_steps else []
                    }
                    st.session_state.all_row_data.append(row_data)
                
                workbook.close()
                st.write(f"✅ Row {row_idx + 1} processed: {mapping_count} fields mapped, {steps_written} steps written -> {filename}")
            
            st.success(f"🎉 Successfully processed {len(data_df)} rows!")
            return True, st.session_state.all_row_data
            
        except Exception as e:
            st.error(f"❌ Error mapping template: {e}")
            st.write("📋 Traceback:", traceback.format_exc())
            return False, []
            
    def map_data_with_section_context_for_row(self, template_fields, data_df, row_idx):
        """Map data for specific row"""
        mapping_results = {}
        used_columns = set()

        try:
            data_columns = data_df.columns.tolist()

            for coord, field in template_fields.items():
                try:
                    best_match = None
                    best_score = 0.0
                    field_value = field['value']
                    section_context = field.get('section_context')

                    # Use existing mapping logic but for specific row
                    if section_context and section_context in self.section_mappings:
                        section_mappings = self.section_mappings[section_context]['field_mappings']

                        for template_field_key, data_column_pattern in section_mappings.items():
                            normalized_field_value = self.preprocess_text(field_value)
                            normalized_template_key = self.preprocess_text(template_field_key)

                            if normalized_field_value == normalized_template_key:
                                if section_context == "procedure_information":
                                    expected_column = data_column_pattern 
                                else:
                                    section_prefix = section_context.split('_')[0].capitalize()
                                    expected_column = f"{section_prefix} {data_column_pattern}".strip()

                                for data_col in data_columns:
                                    if data_col in used_columns:
                                        continue
                                    if self.preprocess_text(data_col) == self.preprocess_text(expected_column):
                                        best_match = data_col
                                        best_score = 1.0
                                        break

                                if not best_match:
                                    for data_col in data_columns:
                                        if data_col in used_columns:
                                            continue
                                        similarity = self.calculate_similarity(expected_column, data_col)
                                        if similarity > best_score and similarity >= self.similarity_threshold:
                                            best_score = similarity
                                            best_match = data_col
                                break

                    # Fallback logic (same as original)
                    if not best_match:
                        for data_col in data_columns:
                            if data_col in used_columns:
                                continue
                            similarity = self.calculate_similarity(field_value, data_col)
                            if similarity > best_score and similarity >= self.similarity_threshold:
                                best_score = similarity
                                best_match = data_col

                    mapping_results[coord] = {
                        'template_field': field_value,
                        'data_column': best_match,
                        'similarity': best_score,
                        'field_info': field,
                        'section_context': section_context,
                        'is_mappable': best_match is not None
                    }

                    if best_match:
                        used_columns.add(best_match)

                except Exception as e:
                    st.error(f"Error mapping field {coord}: {e}")
                    continue

        except Exception as e:
            st.error(f"Error in map_data_with_section_context_for_row: {e}")

        return mapping_results
    
    def write_filled_steps_to_template(self, worksheet, filled_steps):
        """Write filled procedure steps to merged cells B to P starting from Row 28"""
        try:
            from openpyxl.cell import MergedCell
            from openpyxl.styles import Font, Alignment

            print(f"\n=== WRITING FILLED PROCEDURE STEPS ===")
            st.write(f"🔄 Writing {len(filled_steps)} filled procedure steps to template")

            start_row = 28
            target_col = 2  # Column B
            end_col = 18    # Column P

            steps_written = 0

            for i, step in enumerate(filled_steps):
                step_row = start_row + i
                step_text = step.strip()
            
                # Safety check
                if step_row > worksheet.max_row + 20:
                    st.warning(f"⚠️ Stopping at row {step_row} to avoid exceeding template boundaries")
                    break
            
                try:
                    # Define the merge range for this row (B to P)
                    merge_range = f"B{step_row}:R{step_row}"
                    target_cell = worksheet.cell(row=step_row, column=target_col)
                
                    print(f"📝 Writing filled step {i + 1} to {merge_range}: {step_text[:50]}...")
                    st.write(f"📝 Step {i + 1} -> {merge_range}: {step_text[:50]}...")

                    # Unmerge any existing ranges that might conflict
                    existing_merged_ranges = []
                    for merged_range in list(worksheet.merged_cells.ranges):
                        if (merged_range.min_row <= step_row <= merged_range.max_row and
                            merged_range.min_col <= end_col and merged_range.max_col >= target_col):
                            existing_merged_ranges.append(merged_range)

                    for merged_range in existing_merged_ranges:
                        try:
                            worksheet.unmerge_cells(str(merged_range))
                            print(f"🔧 Unmerged existing range: {merged_range}")
                        except Exception as unmerge_error:
                            print(f"⚠️ Warning: Could not unmerge {merged_range}: {unmerge_error}")

                    # Clear any existing content in the range
                    for col in range(target_col, end_col + 1):
                        cell = worksheet.cell(row=step_row, column=col)
                        cell.value = None

                    # Write the step text to the first cell (B)
                    target_cell.value = step_text
                    target_cell.font = Font(name='Calibri', size=10)
                    target_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

                    # Merge the cells B to P for this row
                    try:
                        worksheet.merge_cells(merge_range)
                        print(f"✅ Merged range: {merge_range}")
                    except Exception as merge_error:
                        print(f"⚠️ Warning: Could not merge {merge_range}: {merge_error}")
                        st.warning(f"Could not merge {merge_range}: {merge_error}")
  
                    # Adjust row height based on text length
                    chars_per_line = 120
                    num_lines = max(1, len(step_text) // chars_per_line + 1)
                    estimated_height = 15 + (num_lines - 1) * 15
                    worksheet.row_dimensions[step_row].height = estimated_height

                    steps_written += 1
                
                except Exception as step_error:
                    print(f"❌ Error writing step {i + 1}: {step_error}")
                    st.error(f"Error writing step {i + 1}: {step_error}")
                    continue

            print(f"\n✅ FILLED PROCEDURE STEPS COMPLETED")
            print(f"   Total steps written: {steps_written}")
        
            st.success(f"✅ Successfully wrote {steps_written} filled procedure steps to template")

            return steps_written

        except Exception as e:
            print(f"💥 Critical error in write_filled_steps_to_template: {e}")
            st.error(f"Critical error writing filled procedure steps: {e}")
            return 0

# Packaging types and procedures from reference code
PACKAGING_TYPES = [
    {
        "name": "BOX IN BOX SENSITIVE",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Box%20in%20Box%20sensitive.png",
        "description": "Double protection for sensitive items"
    },
    {
        "name": "BOX IN BOX",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Box%20in%20box.png",
        "description": "Standard double boxing protection"
    },
    {
        "name": "CARTON BOX WITH SEPARATOR FOR ONE PART",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Cardboard%20Box%20with%20Protective%20Packing.png",
        "description": "Single item with internal separator"
    },
    {
        "name": "INDIVIDUAL NOT SENSITIVE",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Individual%20not%20sensitive.png",
        "description": "Individual packaging for standard items"
    },
    {
        "name": "INDIVIDUAL PROTECTION FOR EACH PART MANY TYPE",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Individual%20each%20part%20many%20types.png",
        "description": "Different protection for various parts"
    },
    {
        "name": "INDIVIDUAL PROTECTION FOR EACH PART",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Individual%20for%20each%20part.png",
        "description": "Uniform protection for each part"
    },
    {
        "name": "INDIVIDUAL SENSITIVE",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Individual%20Sensitive.png",
        "description": "Individual packaging for sensitive items"
    },
    {
        "name": "MANY IN ONE TYPE",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/2af9551cb1033072c5d79e029fe17448f8bbc096/Many%20in%20one.png",
        "description": "Multiple items in single packaging"
    },
    {
        "name": "SINGLE BOX",
        "image_url": "https://raw.githubusercontent.com/Rimjhimrani/Pack/88ee0796f874244af8152c681df74d352cf5359a/Single%20Box.png",
        "description": "Simple single box packaging"
    }
]

def display_packaging_grid():
    """Grid-style layout for packaging selection with consistent image sizes"""
    st.header("📦 Step 1: Select Packaging Type")
    st.markdown("Choose the most appropriate packaging type for your needs:")
    
    # Custom CSS for consistent grid image sizing
    st.markdown("""
    <style>
    .grid-image-container {
        display: flex;
        justify-content: center;
        align-items: center;
        height: 200px;
        margin-bottom: 10px;
        border: 1px solid #e0e0e0;
        border-radius: 8px;
        background-color: #fafafa;
    }
    .grid-image-container img {
        max-width: 200px;
        max-height: 200px;
        object-fit: contain;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Create a grid layout for packaging options
    cols = st.columns(3)
    
    for i, packaging in enumerate(PACKAGING_TYPES):
        with cols[i % 3]:
            # Create a container for each packaging option
            with st.container():
                # Display image with consistent sizing
                try:
                    # Use HTML container for consistent sizing
                    st.markdown(f"""
                    <div class="grid-image-container">
                        <img src="{packaging['image_url']}" alt="{packaging['name']}" />
                    </div>
                    """, unsafe_allow_html=True)
                except Exception as e:
                    # Better fallback with consistent sizing
                    st.markdown("""
                    <div class="grid-image-container">
                        <div style="text-align: center; color: #666;">
                            📦<br>Image loading...
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Title and description
                st.markdown(f"**{packaging['name']}**")
                st.write(packaging["description"])
                
                # Selection button with improved styling
                is_selected = st.session_state.get('selected_packaging_type') == packaging['name']
                button_text = "✅ Selected" if is_selected else "Select"
                
                if st.button(
                    button_text,
                    key=f"pkg_{i}", 
                    use_container_width=True,
                    type="primary" if is_selected else "secondary",
                    disabled=is_selected
                ):
                    if not is_selected:
                        st.session_state.selected_packaging_type = packaging['name']
                        st.session_state.selected_packaging_image = packaging['image_url']
                        st.success(f"Selected: {packaging['name']}")
                        navigate_to_step(2)
                        st.rerun()
            
            # Add some spacing
            st.markdown("---")

# Alternative grid method using st.image with fixed parameters
def display_packaging_grid_alternative():
    """Alternative grid-style layout using st.image with fixed dimensions"""
    st.header("📦 Step 1: Select Packaging Type")
    st.markdown("Choose the most appropriate packaging type for your needs:")
    
    # Create a grid layout for packaging options
    cols = st.columns(3)
    
    for i, packaging in enumerate(PACKAGING_TYPES):
        with cols[i % 3]:
            # Create a container for each packaging option
            with st.container():
                # Display image with fixed dimensions for consistency
                try:
                    st.image(
                        packaging["image_url"], 
                        caption=packaging["name"],
                        width=180,  # Fixed width for all images
                        use_container_width=False  # Don't use container width to maintain fixed size
                    )
                except Exception as e:
                    # Consistent fallback
                    st.info("📦 Image loading...")
                    st.write(f"**{packaging['name']}**")
                    st.caption("Image will load shortly...")
                
                # Description
                st.write(packaging["description"])
                
                # Selection button with improved styling
                is_selected = st.session_state.get('selected_packaging_type') == packaging['name']
                button_text = "✅ Selected" if is_selected else "Select"
                
                if st.button(
                    button_text,
                    key=f"pkg_alt_{i}", 
                    use_container_width=True,
                    type="primary" if is_selected else "secondary",
                    disabled=is_selected
                ):
                    if not is_selected:
                        st.session_state.selected_packaging_type = packaging['name']
                        st.session_state.selected_packaging_image = packaging['image_url']
                        st.success(f"Selected: {packaging['name']}")
                        navigate_to_step(2)
                        st.rerun()
            
            # Add some spacing
            st.markdown("---")
            
def main():
    # Header
    st.title("📦 AgiloPACK")
    st.markdown("---")
    
    # Progress indicator
    steps = [
        "Select Packaging Type",
        "Upload Template File", 
        "Upload Data File",
        "Auto-Fill Template",
        "Choose Image Option",
        "Generate Final Document"
    ]

    # Create progress bar
    progress_cols = st.columns(len(steps))
    for i, (col, step) in enumerate(zip(progress_cols, steps)):
        with col:
            if i + 1 < st.session_state.current_step:
                st.success(f"✅ {i+1}. {step}")
            elif i + 1 == st.session_state.current_step:
                st.info(f"🔄 {i+1}. {step}")
            else:
                st.write(f"⏳ {i+1}. {step}")
    
    st.markdown("---")
    
    # Step 1: Select Packaging Type
    if st.session_state.current_step == 1:
        # Display the grid layout directly
        display_packaging_grid()  # Uses HTML/CSS for consistent sizing
        # display_packaging_grid_alternative()  # Alternative: Uses st.image with fixed width
        
        # Show selected packaging details
        if st.session_state.get('selected_packaging_type'):
            st.markdown("### 📋 Selection Summary")
            with st.expander("✅ Selected Packaging Details", expanded=True):
                col1, col2 = st.columns([1, 2])
                with col1:
                    if st.session_state.get('selected_packaging_image'):
                        try:
                            st.image(st.session_state.selected_packaging_image, width=200)
                        except:
                            st.info("📦 Selected Package")
                with col2:
                    st.write(f"**Type:** {st.session_state.selected_packaging_type}")
                    # Find description
                    selected_pkg = next((pkg for pkg in PACKAGING_TYPES if pkg['name'] == st.session_state.selected_packaging_type), None)
                    if selected_pkg:
                        st.write(f"**Description:** {selected_pkg['description']}")
                    
                    # Action buttons
                    col2a, col2b = st.columns(2)
                    with col2a:
                        if st.button("🔄 Change Selection", type="secondary"):
                            st.session_state.selected_packaging_type = None
                            st.session_state.selected_packaging_image = None
                            st.rerun()
                    with col2b:
                        if st.button("Continue to Step 2 →", type="primary"):
                            navigate_to_step(2)
                            st.rerun()
        
    # Step 2: Upload Template File
    elif st.session_state.current_step == 2:
        st.header("📄 Step 2: Upload Template File")
        
        st.info(f"Selected Packaging Type: {st.session_state.selected_packaging_type}")
        
        uploaded_template = st.file_uploader(
            "Choose template file (Excel or Word)",
            type=['xlsx', 'xls', 'docx'],
            key="template_upload"
        )
        
        if uploaded_template is not None:
            # Save uploaded file
            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{uploaded_template.name.split('.')[-1]}") as tmp_file:
                tmp_file.write(uploaded_template.getvalue())
                st.session_state.template_file = tmp_file.name
            
            st.success("✅ Template file uploaded successfully!")
            
            # NEW: Analyze template structure and show preview
            with st.expander("📖 Template Analysis", expanded=True):
                try:
                    mapper = EnhancedTemplateMapperWithImages()
                    
                    # Read procedure steps from template
                    template_procedure_steps = mapper.read_procedure_steps_from_template(
                        st.session_state.template_file,
                        st.session_state.selected_packaging_type
                    )
                    
                    if template_procedure_steps:
                        st.success(f"✅ Found {len(template_procedure_steps)} procedure steps in template")
                        with st.expander("Preview Procedure Steps"):
                            for i, step in enumerate(template_procedure_steps[:5], 1):  # Show first 5
                                st.write(f"{i}. {step[:100]}..." if len(step) > 100 else f"{i}. {step}")
                            if len(template_procedure_steps) > 5:
                                st.write(f"... and {len(template_procedure_steps) - 5} more steps")
                    else:
                        st.warning("⚠️ No procedure steps found in template")
                    
                    # Find template fields
                    template_fields, image_areas = mapper.find_template_fields_with_context_and_images(
                        st.session_state.template_file
                    )
                    
                    if template_fields:
                        st.success(f"✅ Found {len(template_fields)} mappable fields in template")
                        with st.expander("Preview Template Fields"):
                            fields_by_section = {}
                            for field_info in template_fields.values():
                                section = field_info.get('section_context', 'unknown')
                                if section not in fields_by_section:
                                    fields_by_section[section] = []
                                fields_by_section[section].append(field_info['value'])
                            
                            for section, fields in fields_by_section.items():
                                st.write(f"**{section.replace('_', ' ').title()}**: {', '.join(fields[:5])}")
                                if len(fields) > 5:
                                    st.caption(f"... and {len(fields) - 5} more fields")
                    else:
                        st.warning("⚠️ No mappable fields found in template")
                
                except Exception as e:
                    st.error(f"Error analyzing template: {e}")
            
            if st.button("Continue to Data Upload", key="continue_to_step3"):
                navigate_to_step(3)
        
        # Back navigation
        if st.button("← Go Back", key="back_from_2"):
            navigate_to_step(1)
    
    # Step 3: Upload Data File
    elif st.session_state.current_step == 3:
        st.header("📊 Step 3: Upload Data File (Excel)")
        
        uploaded_data = st.file_uploader(
            "Choose Excel data file",
            type=['xlsx', 'xls'],
            key="data_upload"
        )
        
        if uploaded_data is not None:
            # Save uploaded file
            with tempfile.NamedTemporaryFile(delete=False, suffix=f".{uploaded_data.name.split('.')[-1]}") as tmp_file:
                tmp_file.write(uploaded_data.getvalue())
                st.session_state.data_file = tmp_file.name
            
            st.success("✅ Data file uploaded successfully!")
            
            try:
                df = pd.read_excel(st.session_state.data_file)
                st.write("Data Preview:")
                st.dataframe(df.head())
                
                # NEW: Show column analysis for critical fields
                with st.expander("📊 Data Column Analysis"):
                    critical_fields = {
                        "Procedure Fields": ["Layer", "Level", "x No. of Parts", "x No of Parts"],
                        "Inner Dimensions": ["Inner L", "Inner W", "Inner H", "Inner Qty/Pack"],
                        "Outer Dimensions": ["Outer L", "Outer W", "Outer H"],
                        "Primary Packaging": ["Primary L-mm", "Primary W-mm", "Primary H-mm", "Primary Qty/Pack"],
                        "Secondary Packaging": ["Secondary L-mm", "Secondary W-mm", "Secondary H-mm"],
                        "Part Information": ["Part No", "Part Description", "Vendor Code", "Vendor Name"]
                    }
                    
                    found_fields = {}
                    for category, fields in critical_fields.items():
                        found_fields[category] = []
                        for field in fields:
                            matching_cols = [col for col in df.columns if field.lower() in col.lower()]
                            if matching_cols:
                                found_fields[category].extend(matching_cols)
                    
                    for category, fields in found_fields.items():
                        if fields:
                            st.success(f"✅ **{category}**: {', '.join(fields)}")
                        else:
                            st.warning(f"⚠️ **{category}**: No matching columns found")
                
            except Exception as e:
                st.error(f"Error reading data file: {e}")
            
            if st.button("Continue to Auto-Fill", key="continue_to_step4"):
                navigate_to_step(4)
        
        # Back navigation
        if st.button("← Go Back", key="back_from_3"):
            navigate_to_step(2)
    
    # Step 4: Auto-Fill Template (ENHANCED)
    elif st.session_state.current_step == 4:
        st.header("🔄 Step 4: Auto-Fill Template with Enhanced Processing")
    
        if st.session_state.mapping_completed and hasattr(st.session_state, 'all_row_data'):
            st.success(f"✅ Template auto-filling completed for {len(st.session_state.all_row_data)} rows!")
        
            # Show enhanced summary of processed rows
            with st.expander("View Enhanced Processing Summary", expanded=True):
                total_mappings = sum(row['mapping_count'] for row in st.session_state.all_row_data)
                total_steps = sum(row['steps_written'] for row in st.session_state.all_row_data)
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("📄 Templates Generated", len(st.session_state.all_row_data))
                with col2:
                    st.metric("🔗 Total Field Mappings", total_mappings)
                with col3:
                    st.metric("📝 Procedure Steps Written", total_steps)
                
                # Detailed row information
                for i, row_data in enumerate(st.session_state.all_row_data):
                    with st.container():
                        col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
                        with col1:
                            st.write(f"**Row {i+1}**: {row_data['filename']}")
                        with col2:
                            st.write(f"📊 {row_data['mapping_count']} mappings")
                        with col3:
                            st.write(f"📝 {row_data['steps_written']} steps")
                        with col4:
                            part_no = row_data.get('part_no', 'N/A')
                            st.write(f"🔧 {part_no}")
        
            if st.button("Continue to Image Options", key="continue_to_images"):
                navigate_to_step(5)
    
        else:
            # Enhanced auto-fill process
            st.info("🚀 Enhanced auto-fill will process templates with procedure steps from your template")
            
            # Show what will be processed
            if st.session_state.template_file and st.session_state.data_file:
                try:
                    # Preview what will be processed
                    df_preview = pd.read_excel(st.session_state.data_file)
                    st.write(f"📊 Ready to process: {len(df_preview)} rows of data")
                    
                    mapper = EnhancedTemplateMapperWithImages()
                    
                    # Show procedure steps that will be used
                    template_steps = mapper.read_procedure_steps_from_template(st.session_state.template_file)
                    if template_steps:
                        st.info(f"📖 Will use {len(template_steps)} procedure steps from your template")
                        with st.expander("Preview Procedure Steps Processing"):
                            st.write("**Template contains these procedure steps with placeholders:**")
                            for i, step in enumerate(template_steps[:3], 1):
                                st.write(f"{i}. {step}")
                            if len(template_steps) > 3:
                                st.write(f"... and {len(template_steps) - 3} more steps")
                            
                            st.write("**These placeholders will be replaced with actual data:**")
                            placeholders = ["{x No. of Parts}", "{Level}", "{Layer}", "{Inner L}", "{Inner W}", "{Inner H}", 
                                          "{Outer L}", "{Outer W}", "{Outer H}", "{Primary Qty/Pack}", "{Secondary L-mm}"]
                            st.write(", ".join(placeholders))
                    else:
                        st.warning("⚠️ No procedure steps found in template")
                    
                except Exception as e:
                    st.error(f"Error previewing processing: {e}")
            
            if st.button("🚀 Start Enhanced Auto-Fill Process", key="start_enhanced_autofill", type="primary"):
                with st.spinner("🔄 Processing templates with enhanced procedure step handling..."):
                    try:
                        # Create progress tracking
                        progress_container = st.container()
                        status_container = st.container()
                        
                        mapper = EnhancedTemplateMapperWithImages()
                        
                        # Enhanced mapping with procedure steps
                        success, all_row_data = mapper.map_template_with_data(
                            st.session_state.template_file,
                            st.session_state.data_file
                        )
                        
                        if success and all_row_data:
                            st.session_state.mapping_completed = True
                            st.session_state.all_row_data = all_row_data
                            
                            # Show success summary
                            with status_container:
                                st.success("🎉 Enhanced auto-fill completed successfully!")
                                
                                # Enhanced metrics
                                total_mappings = sum(row['mapping_count'] for row in all_row_data)
                                total_steps = sum(row['steps_written'] for row in all_row_data)
                                rows_with_steps = sum(1 for row in all_row_data if row['steps_written'] > 0)
                                
                                col1, col2, col3, col4 = st.columns(4)
                                with col1:
                                    st.metric("📄 Templates", len(all_row_data))
                                with col2:
                                    st.metric("🔗 Field Mappings", total_mappings)
                                with col3:
                                    st.metric("📝 Procedure Steps", total_steps)
                                with col4:
                                    st.metric("✅ Success Rate", f"{rows_with_steps}/{len(all_row_data)}")
                            
                            st.rerun()
                        else:
                            st.error("❌ Enhanced auto-fill process failed")
                        
                    except Exception as e:
                        st.error(f"❌ Error during enhanced auto-fill: {e}")
                        st.write("**Error Details:**")
                        st.code(traceback.format_exc())
    
        if st.button("← Go Back", key="back_from_4"):
            navigate_to_step(3)
    
    # Step 5: Choose Image Option (SAME AS BEFORE)
    elif st.session_state.current_step == 5:
        st.header("🖼️ Step 5: Choose Image Option")

        # Three column layout for three options
        col1, col2, col3 = st.columns(3)

        with col1:
            if st.button("🔍 Smart Extract from Data File", use_container_width=True):
                st.session_state.image_option = 'extract'
        
                # Enhanced image extraction
                with st.spinner("🧠 Analyzing and extracting images..."):
                    extractor = EnhancedImageExtractor()  # Use the new enhanced extractor
                    extracted_images = extractor.extract_images_from_excel(st.session_state.data_file)
            
                    if extracted_images and 'all_sheets' in extracted_images:
                        st.session_state.extracted_excel_images = extracted_images['all_sheets']
                        st.success(f"✅ Intelligently extracted {len(st.session_state.extracted_excel_images)} images!")
                
                        # Enhanced preview with grouping
                        st.write("**📊 Extracted Images Analysis:**")
                        image_types = {}
                        for img_key, img_data in st.session_state.extracted_excel_images.items():
                            img_type = img_data['type']
                            if img_type not in image_types:
                                image_types[img_type] = 0
                            image_types[img_type] += 1
                
                        # Show type distribution
                        cols = st.columns(len(image_types))
                        for i, (img_type, count) in enumerate(image_types.items()):
                            with cols[i]:
                                st.metric(f"{img_type.capitalize()}", count)
                
                        # Show confidence levels
                        high_confidence = sum(1 for img in st.session_state.extracted_excel_images.values() 
                                              if img.get('confidence', 0) > 0.7)
                        st.info(f"🎯 {high_confidence} images classified with high confidence")
                
                    else:
                        st.warning("No images found in the Excel file")

        with col2:
            if st.button("📁 Upload New Images", use_container_width=True):
                st.session_state.image_option = 'upload'

        with col3:
            if st.button("📄 Generate Without Images", use_container_width=True):
                st.session_state.image_option = 'no_images'

        # Handle upload new images option (Enhanced)
        if st.session_state.image_option == 'upload':
            st.subheader("📤 Upload Images by Type")
    
            image_types = ['current', 'primary', 'secondary', 'label']
            type_descriptions = {
                'current': 'Current/Present packaging state',
                'primary': 'Inner/Primary packaging',
                'secondary': 'Outer/Secondary packaging', 
                'label': 'Labels, barcodes, or identification'
            }
    
            uploaded_count = 0
            for img_type in image_types:
                with st.expander(f"📋 {img_type.capitalize()} Packaging Image", expanded=False):
                    st.write(f"*{type_descriptions[img_type]}*")
            
                    uploaded_img = st.file_uploader(
                        f"Choose {img_type} image",
                        type=['png', 'jpg', 'jpeg', 'gif', 'bmp'],
                        key=f"img_upload_{img_type}"
                    )
            
                    if uploaded_img is not None:
                        # Convert to base64
                        img_bytes = uploaded_img.read()
                        img_b64 = base64.b64encode(img_bytes).decode()
                
                        # Store in session state with enhanced metadata
                        st.session_state.uploaded_images[f"{img_type}_uploaded"] = {
                            'data': img_b64,
                            'format': uploaded_img.type.split('/')[-1].upper(),
                            'size': (0, 0),  # Will be determined when processing
                            'type': img_type,
                            'confidence': 1.0,  # User uploaded = high confidence
                            'source': 'user_upload'
                        }
                
                        uploaded_count += 1
                
                        # Preview with analysis
                        col1, col2 = st.columns([1, 2])
                        with col1:
                            st.image(f"data:image/{uploaded_img.type.split('/')[-1]};base64,{img_b64}", width=150)
                        with col2:
                            st.success(f"✅ {img_type.capitalize()} image uploaded")
                            st.write(f"**Size**: {len(img_bytes):,} bytes")
                            st.write(f"**Format**: {uploaded_img.type}")
    
            if uploaded_count > 0:
                st.success(f"📁 {uploaded_count} images uploaded successfully!")

        # Handle no images option
        elif st.session_state.image_option == 'no_images':
            st.subheader("📄 Text-Only Generation Mode")
            st.info("""
                **📝 Text-Only Mode Selected**
                Your documents will be generated using only the data from your spreadsheet without any images. This mode:
                ✅ **Faster Processing** - Quicker generation without image analysis  
                ✅ **Smaller File Sizes** - Lighter documents for easier sharing  
                ✅ **Focus on Content** - Emphasizes textual information and data  
                ✅ **Universal Compatibility** - Works with all systems and formats  
                **Note:** You can always regenerate with images later if needed.
            """)
        
            # Optional: Allow users to add a note about why no images
            image_note = st.text_area(
                "📝 Optional: Add a note about image availability",
                placeholder="e.g., 'Images to be provided separately' or 'Product images pending approval'",
                help="This note will be included in generated documents to explain the absence of images"
            )
        
            if image_note:
                st.session_state.no_images_note = image_note
        
            # Show what will be included instead
            with st.expander("📋 What will be included in text-only mode"):
                st.write("""
                    - ✅ All part numbers and specifications
                    - ✅ Vendor information and codes  
                    - ✅ Descriptions and technical details
                    - ✅ Packaging requirements and notes
                    - ✅ Regulatory and compliance information
                    - ✅ Custom formatting and styling
                    - ❌ Product images and visual references
                    - ❌ Packaging photos and diagrams
                """)

        # Auto-match option for extracted images (only show if extract option is selected)
        if (st.session_state.image_option == 'extract' and st.session_state.extracted_excel_images and hasattr(st.session_state, 'all_row_data')):
            st.subheader("🎯 Auto-Match Images to Parts")
            if st.button("🤖 Auto-Match Images to Current Part Data"):
                with st.spinner("Matching images to part numbers..."):
                    matched_results = {}
                    extractor = EnhancedImageExtractor()
                    # Store the extracted images in the correct format expected by the method
                    all_extracted_images = {'all_sheets': st.session_state.extracted_excel_images}
                    for idx, row_data in enumerate(st.session_state.all_row_data):
                        part_no = row_data.get('part_no', f'Part_{idx}')
                        vendor_code = row_data.get('vendor_code', 'Unknown')
                        st.write(f"🔍 Processing part: {part_no}")
                        # Correct method call with proper parameters
                        part_images = extractor.extract_images_for_part(
                            st.session_state.data_file,
                            part_no,
                            all_extracted_images,  # This was missing!
                            vendor_code,
                            current_row=idx + 2  # Assuming row 1 is header, so data starts from row 2
                        )
            
                        if part_images:
                            matched_results[row_data.get('filename', f'Part_{idx}')] = {
                                'images': part_images,
                                'count': len(part_images),
                                'part_no': part_no,
                                'vendor': vendor_code
                            }
                            st.write(f"✅ Found {len(part_images)} images for {part_no}")
                        else:
                            st.write(f"⚠️ No images found for {part_no}")
                            matched_results[row_data.get('filename', f'Part_{idx}')] = {
                                'images': {},
                                'count': 0,
                                'part_no': part_no,
                                'vendor': vendor_code
                            }
                    st.session_state.matched_part_images = matched_results
                    # Show matching summary
                    total_matched = sum(result['count'] for result in matched_results.values())
                    parts_with_images = len([r for r in matched_results.values() if r['count'] > 0])
        
                    st.success(f"🎯 Matched {total_matched} images across {parts_with_images}/{len(matched_results)} parts")
        
                    # Detailed breakdown
                    with st.expander("📊 Matching Details"):
                        for filename, result in matched_results.items():
                            if result['count'] > 0:
                                st.success(f"**{result['part_no']}** ({result['vendor']}): {result['count']} images")
                                # Show image types found
                                image_types = {}
                                for img_key, img_data in result['images'].items():
                                    img_type = img_data.get('type', 'unknown')
                                    image_types[img_type] = image_types.get(img_type, 0) + 1
                                type_summary = ", ".join([f"{t}: {c}" for t, c in image_types.items()])
                                st.caption(f"Types found: {type_summary}")
                            else:
                                st.warning(f"**{result['part_no']}** ({result['vendor']}): No images")

        # Continue button with enhanced validation
        can_continue = False
        if st.session_state.image_option == 'extract':
            can_continue = (st.session_state.extracted_excel_images or hasattr(st.session_state, 'matched_part_images'))
        elif st.session_state.image_option == 'upload':
            can_continue = len(st.session_state.uploaded_images) > 0
        elif st.session_state.image_option == 'no_images':
            can_continue = True  # No images option always allows continuation

        if can_continue:
            # Show different button text based on option
            button_text = "🚀 Continue to Final Generation"
            if st.session_state.image_option == 'no_images':
                button_text = "📄 Continue to Text-Only Generation"
        
            if st.button(button_text, key="continue_to_step6", type="primary"):
                navigate_to_step(6)
        else:
            if st.session_state.image_option == 'extract':
                st.info("📋 Please extract images or match images to parts before continuing")
            elif st.session_state.image_option == 'upload':
                st.info("📋 Please upload at least one image before continuing")
            else:
                st.info("📋 Please select an image option to continue")
    
        # Back navigation
        if st.button("← Go Back", key="back_from_5"):
            navigate_to_step(4)

    # Step 6: Generate Final Document (SAME AS BEFORE - keeping your existing complex logic)
    elif st.session_state.current_step == 6:
        st.header("🎨 Step 6: Generate Final Documents with Smart Placement")
    
        # Configuration options
        with st.expander("⚙️ Generation Settings", expanded=True):
            col1, col2 = st.columns(2)
            with col1:
                use_smart_placement = st.checkbox("🧠 Use Smart Image Placement", value=True)
                preserve_aspect_ratio = st.checkbox("📐 Preserve Image Aspect Ratios", value=True)
            with col2:
                image_quality = st.selectbox("🖼️ Image Quality", ["High", "Medium", "Low"], index=0)
                max_image_size = st.slider("📏 Max Image Size (cm)", 2, 10, 5)
    
        # Generation summary
        total_templates = len(st.session_state.all_row_data) if hasattr(st.session_state, 'all_row_data') else 0
        total_images = 0
    
        if st.session_state.image_option == 'extract':
            if hasattr(st.session_state, 'matched_part_images'):
                total_images = sum(result['count'] for result in st.session_state.matched_part_images.values())
            else:
                total_images = len(st.session_state.extracted_excel_images)
        elif st.session_state.image_option == 'upload':
            total_images = len(st.session_state.uploaded_images)
    
        # Summary metrics
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("📄 Templates to Generate", total_templates)
        with col2:
            st.metric("🖼️ Total Images Available", total_images)
        with col3:
            st.metric("🎯 Placement Mode", "Smart" if use_smart_placement else "Fixed")
    
        if st.button("🚀 Generate All Templates with Enhanced Placement", type="primary", use_container_width=True):
            with st.spinner("🎨 Generating templates with smart image placement..."):
                try:
                    extractor = EnhancedImageExtractor()
                    generated_files = []
                    generation_log = []
                
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                
                    for i, row_data in enumerate(st.session_state.all_row_data):
                        status_text.text(f"Processing template {i+1}/{total_templates}: {row_data['filename']}")
                        progress_bar.progress((i + 1) / total_templates)
                    
                        # Load the mapped template for this row
                        workbook = openpyxl.load_workbook(row_data['file_path'])
                        worksheet = workbook.active
                    
                        # Determine images to add
                        images_to_add = {}
                    
                        if st.session_state.image_option == 'extract':
                            if hasattr(st.session_state, 'matched_part_images'):
                                # Use pre-matched part-specific images
                                matched_data = st.session_state.matched_part_images.get(row_data['filename'], {})
                                images_to_add = matched_data.get('images', {})
                            else:
                                # Extract images specific to this part on-the-fly
                                images_to_add = extractor.extract_images_for_part(
                                    st.session_state.data_file,
                                    row_data.get('part_no', ''),
                                    row_data.get('description', ''),
                                    row_data.get('vendor_code', '')
                                )
                        elif st.session_state.image_option == 'upload':
                            # Use same uploaded images for all templates
                            images_to_add = st.session_state.uploaded_images
                    
                        # Add images to template using enhanced placement
                        if images_to_add:
                            if use_smart_placement:
                                added_count, temp_paths = extractor.smart_add_images_to_template(
                                    st.session_state.template_file, worksheet, images_to_add
                                )
                                placement_method = "Smart Analysis"
                            else:
                                added_count, temp_paths = extractor.add_images_to_template(
                                    worksheet, images_to_add
                                )
                                placement_method = "Fixed Positions"
                        
                            generation_log.append({
                                'template': row_data['filename'],
                                'images_added': added_count,
                                'placement_method': placement_method,
                                'part_no': row_data.get('part_no', 'N/A'),
                                'vendor': row_data.get('vendor_code', 'N/A')
                            })
                        else:
                            generation_log.append({
                                'template': row_data['filename'],
                                'images_added': 0,
                                'placement_method': 'No Images',
                                'part_no': row_data.get('part_no', 'N/A'),
                                'vendor': row_data.get('vendor_code', 'N/A')
                            })
                    
                        # Save final document with enhanced naming
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        part_no_safe = re.sub(r'[^\w\-_.]', '_', str(row_data.get('part_no', 'Unknown')))
                        vendor_safe = re.sub(r'[^\w\-_.]', '_', str(row_data.get('vendor_code', 'Unknown')))
                        
                        # Add row index to ensure uniqueness - this prevents duplicate filenames
                        row_index = str(i + 1).zfill(3)  # Zero-padded to 3 digits (001, 002, etc.)
                        final_filename = f"{vendor_safe}_{part_no_safe}_R{row_index}_{timestamp}.xlsx"
                    
                        final_filename = f"{vendor_safe}_{part_no_safe}_{timestamp}.xlsx"
                    
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                            workbook.save(tmp_file.name)
                        
                            with open(tmp_file.name, 'rb') as f:
                                file_bytes = f.read()
                        
                            generated_files.append({
                                'filename': final_filename,
                                'data': file_bytes,
                                'row_info': row_data,
                                'images_count': len(images_to_add),
                                'generation_info': generation_log[-1]
                            })
                    
                        workbook.close()
                    
                        # Cleanup temporary image files
                        if 'temp_paths' in locals():
                            for temp_path in temp_paths:
                                try:
                                    os.unlink(temp_path)
                                except:
                                    pass
                
                    progress_bar.progress(1.0)
                    status_text.text("✅ Generation complete!")
                
                    st.success(f"🎉 Successfully generated {len(generated_files)} enhanced templates!")
                
                    # Show generation summary
                    st.subheader("📊 Generation Summary")
                
                    summary_df = pd.DataFrame(generation_log)
                    col1, col2 = st.columns([2, 1])
                
                    with col1:
                        st.dataframe(summary_df, use_container_width=True)
                
                    with col2:
                        total_images_placed = summary_df['images_added'].sum()
                        templates_with_images = len(summary_df[summary_df['images_added'] > 0])
                    
                        st.metric("🖼️ Total Images Placed", total_images_placed)
                        st.metric("📄 Templates with Images", templates_with_images)
                        st.metric("📈 Success Rate", f"{templates_with_images/len(summary_df)*100:.1f}%")
                
                    # Enhanced download section
                    st.subheader("📥 Download Generated Templates")
                
                    # Tabs for different download options
                    tab1, tab2, tab3 = st.tabs(["📋 Individual Downloads", "📦 Bulk Download", "📊 Generation Report"])
                
                    with tab1:
                        for file_idx, file_info in enumerate(generated_files):
                            with st.container():
                                col1, col2, col3 = st.columns([2, 1, 1])
                                with col1:
                                    st.write(f"**{file_info['filename']}**")
                                    st.caption(f"Vendor: {file_info['row_info'].get('vendor_code', 'N/A')} | "
                                               f"Part: {file_info['row_info'].get('part_no', 'N/A')} | "
                                               f"Images: {file_info['images_count']}")
                                with col2:
                                    st.write(f"📊 {file_info['generation_info']['placement_method']}")
                                    if file_info['images_count'] > 0:
                                        st.success(f"✅ {file_info['images_count']} images")
                                    else:
                                        st.warning("⚠️ No images")
                                with col3:
                                    # FIXED: Use file index to ensure unique keys
                                    st.download_button(
                                        label="📥 Download",
                                        data=file_info['data'],
                                        file_name=file_info['filename'],
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"download_file_{file_idx}_{timestamp}"  # Unique key using index and timestamp
                                    )
                                st.divider()
                
                    with tab2:
                        if len(generated_files) > 1:
                            # Create ZIP with organized structure
                            zip_buffer = io.BytesIO()
                        
                            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                                # Organize by vendor
                                vendor_folders = {}
                                for file_info in generated_files:
                                    vendor = file_info['row_info'].get('vendor_code', 'Unknown_Vendor')
                                    vendor_safe = re.sub(r'[^\w\-_.]', '_', vendor)
                                
                                    if vendor_safe not in vendor_folders:
                                        vendor_folders[vendor_safe] = []
                                    vendor_folders[vendor_safe].append(file_info)
                            
                                # Add files organized by vendor
                                for vendor_folder, files in vendor_folders.items():
                                    for file_info in files:
                                        zip_path = f"{vendor_folder}/{file_info['filename']}"
                                        zip_file.writestr(zip_path, file_info['data'])
                            
                                # Add generation report
                                report_content = "Template Generation Report\n"
                                report_content += "=" * 40 + "\n\n"
                                report_content += f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                                report_content += f"Total templates: {len(generated_files)}\n"
                                report_content += f"Total images placed: {total_images_placed}\n"
                                report_content += f"Placement method: {'Smart Analysis' if use_smart_placement else 'Fixed Positions'}\n\n"
                            
                                report_content += "Individual Template Details:\n"
                                report_content += "-" * 30 + "\n"
                                for log_entry in generation_log:
                                    report_content += f"Template: {log_entry['template']}\n"
                                    report_content += f"  Part No: {log_entry['part_no']}\n"
                                    report_content += f"  Vendor: {log_entry['vendor']}\n"
                                    report_content += f"  Images Added: {log_entry['images_added']}\n"
                                    report_content += f"  Method: {log_entry['placement_method']}\n\n"
                            
                                zip_file.writestr("Generation_Report.txt", report_content)
                        
                            zip_buffer.seek(0)
                        
                            col1, col2 = st.columns(2)
                            with col1:
                                st.info("📁 Organized by vendor folders with generation report included")
                        
                            with col2:
                                st.download_button(
                                    label="📦 Download All Templates (ZIP)",
                                    data=zip_buffer.getvalue(),
                                    file_name=f"Enhanced_Templates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                                    mime="application/zip",
                                    key="download_enhanced_zip",
                                    use_container_width=True
                                )
                
                    with tab3:
                        # Detailed generation report
                        st.write("**📈 Performance Metrics**")
                    
                        metrics_col1, metrics_col2, metrics_col3 = st.columns(3)
                        with metrics_col1:
                            avg_images_per_template = total_images_placed / len(generated_files) if generated_files else 0
                            st.metric("📊 Avg Images per Template", f"{avg_images_per_template:.1f}")
                    
                        with metrics_col2:
                            placement_methods = summary_df['placement_method'].value_counts()
                            most_common_method = placement_methods.index[0] if not placement_methods.empty else "None"
                            st.metric("🎯 Primary Placement Method", most_common_method)
                    
                        with metrics_col3:
                            templates_without_images = len(summary_df[summary_df['images_added'] == 0])
                            st.metric("⚠️ Templates Without Images", templates_without_images)
                    
                        # Detailed breakdown by vendor
                        if len(summary_df) > 0:
                            st.write("**📊 Breakdown by Vendor**")
                            vendor_summary = summary_df.groupby('vendor').agg({
                                'images_added': ['count', 'sum', 'mean'],
                                'template': 'count'
                            }).round(1)
                        
                            vendor_summary.columns = ['Templates', 'Total Images', 'Avg Images', 'Template Count']
                            st.dataframe(vendor_summary, use_container_width=True)
            
                except Exception as e:
                    st.error(f"❌ Error generating enhanced templates: {e}")
                    st.write("**Error Details:**")
                    st.code(traceback.format_exc())
                
                    # Provide diagnostic information
                    st.write("**Diagnostic Information:**")
                    st.write(f"- Template file exists: {os.path.exists(st.session_state.template_file) if hasattr(st.session_state, 'template_file') else 'No template file'}")
                    st.write(f"- Data file exists: {os.path.exists(st.session_state.data_file) if hasattr(st.session_state, 'data_file') else 'No data file'}")
                    st.write(f"- Row data available: {len(st.session_state.all_row_data) if hasattr(st.session_state, 'all_row_data') else 0} rows")
                    st.write(f"- Image option: {st.session_state.image_option}")
    
        # Advanced options
        with st.expander("🛠️ Advanced Options"):
            col1, col2 = st.columns(2)
        
            with col1:
                if st.button("🔍 Analyze Template Structure"):
                    if hasattr(st.session_state, 'template_file'):
                        with st.spinner("Analyzing template..."):
                            extractor = EnhancedImageExtractor()
                            zones = extractor.analyze_template_structure(st.session_state.template_file)
                        
                            st.write("**Detected Image Placement Zones:**")
                            for zone_type, zone_info in zones.items():
                                if zone_info:
                                    st.success(f"✅ {zone_type}: {zone_info['cell']} "
                                               f"({zone_info['width_cells']}x{zone_info['height_cells']} cells)")
                                else:
                                    st.warning(f"⚠️ {zone_type}: No suitable zone found")
        
            with col2:
                if st.button("📊 Preview Image Assignments") and st.session_state.image_option == 'extract':
                    if hasattr(st.session_state, 'matched_part_images'):
                        st.write("**Image-to-Part Assignments:**")
                        for filename, result in st.session_state.matched_part_images.items():
                            with st.expander(f"{filename} ({result['count']} images)"):
                                for img_key, img_data in result['images'].items():
                                    col1, col2 = st.columns([1, 2])
                                    with col1:
                                        st.image(f"data:image/png;base64,{img_data['data']}", width=100)
                                    with col2:
                                        st.write(f"**Type**: {img_data['type']}")
                                        st.write(f"**Confidence**: {img_data.get('confidence', 0.5):.1f}")
                                        st.write(f"**Size**: {img_data['size']}")
    
        # Back navigation
        if st.button("← Go Back", key="back_from_6"):
            navigate_to_step(5)
    
    # Sidebar with help and information
    with st.sidebar:
        st.header("ℹ️ Help & Information")
        
        st.subheader("Current Progress")
        st.write(f"**Step**: {st.session_state.current_step}/6")
        if st.session_state.selected_packaging_type:
            st.write(f"**Packaging Type**: {st.session_state.selected_packaging_type}")
        
        # Enhanced status display
        if hasattr(st.session_state, 'all_row_data') and st.session_state.all_row_data:
            st.subheader("📊 Processing Status")
            total_templates = len(st.session_state.all_row_data)
            total_mappings = sum(row['mapping_count'] for row in st.session_state.all_row_data)
            total_steps = sum(row['steps_written'] for row in st.session_state.all_row_data)
            
            st.write(f"**Templates Ready**: {total_templates}")
            st.write(f"**Field Mappings**: {total_mappings}")
            st.write(f"**Procedure Steps**: {total_steps}")
        
        st.subheader("Instructions")
        st.write("""
        1. **Select Packaging Type**: Choose from predefined packaging types
        2. **Upload Template**: Upload your Excel template file
        3. **Upload Data**: Upload Excel file with part data
        4. **Auto-Fill**: Enhanced AI mapping with procedure steps
        5. **Add Images**: Extract from Excel or upload new images
        6. **Generate**: Create final template with images
        """)
        
        st.subheader("✨ Enhanced Features")
        st.write("""
        - **Smart Procedure Processing**: Reads steps directly from template
        - **Placeholder Substitution**: Replaces {placeholders} with real data
        - **Section-Based Mapping**: Intelligent field recognition
        - **Multi-Row Processing**: Handles multiple parts automatically
        """)
        
        st.subheader("Supported Formats")
        st.write("**Template Files**: .xlsx, .xls, .docx")
        st.write("**Data Files**: .xlsx, .xls")
        st.write("**Image Files**: .png, .jpg, .jpeg, .gif, .bmp")
        
        # Enhanced reset with confirmation
        if st.button("🔄 Reset All", type="secondary"):
            # Show confirmation in main area
            st.session_state.show_reset_confirmation = True
            st.rerun()
        
        # Handle reset confirmation
        if st.session_state.get('show_reset_confirmation', False):
            st.warning("⚠️ This will clear all progress and uploaded files. Are you sure?")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("✅ Yes, Reset", type="primary"):
                    # Clean up temporary files
                    if hasattr(st.session_state, 'all_row_data'):
                        for row_data in st.session_state.all_row_data:
                            try:
                                if 'file_path' in row_data and os.path.exists(row_data['file_path']):
                                    os.unlink(row_data['file_path'])
                            except:
                                pass
                    
                    # Reset session state
                    for key in list(st.session_state.keys()):
                        if key != 'current_step':
                            del st.session_state[key]
                    st.session_state.current_step = 1
                    st.rerun()
            with col2:
                if st.button("❌ Cancel"):
                    st.session_state.show_reset_confirmation = False
                    st.rerun()


if __name__ == "__main__":
    main()
